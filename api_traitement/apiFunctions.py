import datetime
# from time import gmtime, strftime, strptime
import os
import pandas as pd
import openpyxl as op
import numpy as np
# import requests
import json

from api_traitement.json_fonctions import *

from webapps.models import User
from django.contrib.auth import authenticate


########### Token ###########

# def getToken(baseUrl, data):
#     """
#     Adelphe
#     data = {
#             "config.login": "username",
#             "config.password": "password",
#             "config.databaseName": "database",
#             "referentialLocale": "FR",
#         }
#     """

#     url = baseUrl + "/init/open"
#     response = requests.get(url, params=data)
#     print(response.url)
#     token = response.json()['authenticationToken']

#     return token


# def is_valid(token):
#     """ 
#     Clem
#     Fonction booléenne qui test si le token est encore valide
#     Args:
#         token (str): _description_
#     """
#     api_base = 'https://observe.ob7.ird.fr/observeweb/api/public/init/information?'
#     # Constitution du lien url pour accéder à l'API et fermer la connexion
#     api_url = api_base + 'authenticationToken=' + token
#     response = requests.get(api_url)
#     print("reponse of is valid function ", response.status_code)
#     return response.status_code == 200


# def reload_token(req, username, password):

#     user = authenticate(req, username=username,  password=password)
#     data_user = User.objects.get(username=user)

#     baseUrl = data_user.url

#     print("data_user.database",data_user.database)
#     if data_user.database == 'test' :
#         data_user.username = 'technicienweb'

#     data_user_connect = {
#         "config.login": data_user.username,
#         "config.login": data_user.username,
#         "config.password": password,
#         "config.databaseName": data_user.database,
#         "referentialLocale": data_user.ref_language,
#     }

#     return getToken(baseUrl, data_user_connect)


# def get_all_referential_data(token, module, baseUrl):
#     url = baseUrl + "/referential/" + module + "?authenticationToken=" + token
#     ac_cap = requests.get(url)
#     if ac_cap.status_code == 200:
#         dicoModule = {}
#         for val in json.loads(ac_cap.text)["content"]:
#             vals = val.rsplit('.', 1)[1]
#             dicoModule[vals] = []

#             for valin in json.loads(ac_cap.text)["content"][val]:
#                 dicoModule[vals].append(valin)
#         print("="*20, "get_all_referential_data", "="*20)
#         # print(dicoModule)
#         return dicoModule
#     else:
#         return "Problème de connexion pour recuperer les données"


# # Recuperer les données de la senne en les stockant dans un dossier en local chaque 24
# # En utilisant notre fuiseau horaire
# def load_data(token, baseUrl, forceUpdate=False):
#     print("_"*20, "load_data function starting", "_"*20)
#     day = strftime("%Y-%m-%d", gmtime())

#     # Si les dossiers ne sont pas existant, on les créés
#     if not os.path.exists("media/data"):
#         os.makedirs("media/data")

#     if not os.path.exists("media/temporary_files"):
#         os.makedirs("media/temporary_files")

#     files = os.listdir("media/data")

#     def subFunction(token, day, url):
#         ref_common = get_all_referential_data(token, "common", url)
#         ps_logbook = get_all_referential_data(token, "ps/logbook", url)
#         ps_common = get_all_referential_data(token, "ps/common", url)
#         ll_common = get_all_referential_data(token, "ll/common", url)

#         program = {
#             'Program': {
#                 'seine' :ps_common["Program"],
#                 'longline':ll_common["Program"]
#             }
#         }
#         vesselActivity = {
#             'VesselActivity': {
#                 'seine' :ps_common["VesselActivity"],
#                 'longline':ll_common["VesselActivity"]
#             }
#         }

#         # Suppression des éléments suivant
#         del ps_common["Program"]
#         del ll_common["Program"]
#         del ps_common["VesselActivity"]
#         del ll_common["VesselActivity"]

#         allData = {**ref_common, **ps_logbook, **ps_common, **ll_common, **program, **vesselActivity}
#         # allData = {**ref_common, **ps_logbook, **ps_common}

#         ref_common = get_all_referential_data(token, "common", url)
#         # ref_common2 ="https://observe.ob7.ird.fr/observeweb/api/public/referential/common?authenticationToken=6811592f-bf3b-4fa0-8320-58a4a58c9ab7"
#         ps_logbook = get_all_referential_data(token, "ps/logbook", url)
#         ps_common = get_all_referential_data(token, "ps/common", url)
#         ll_common = get_all_referential_data(token, "ll/common", url)    

#         print("="*20, "load_data SubFunction", "="*20)
#         # print(ref_common[5:])
#         # with open('allData_load.json', 'w', encoding='utf-8') as f:
#         #     json.dump(allData, f, ensure_ascii=False, indent=4)

#         file_name = "media/data/data_" + str(day) + ".json"

#         with open(file_name, 'w', encoding='utf-8') as f:
#             f.write(json.dumps(allData, ensure_ascii=False, indent=4))

#         return allData

#     if (0 < len(files)) and (len(files) <= 1) and (forceUpdate == False):

#         last_date = files[0].split("_")[1].split(".")[0]
#         last_file = files[0]

#         formatted_date1 = strptime(day, "%Y-%m-%d")
#         formatted_date2 = strptime(last_date, "%Y-%m-%d")

#         # Verifier si le jour actuel est superieur au jour precedent
#         if (formatted_date1 > formatted_date2):
#             allData = subFunction(token, day, baseUrl)

#             # Suprimer l'ancienne
#             os.remove("media/data/" + last_file)

#             print("="*20, "allData updated", "="*20)
#             # print(allData[5:])

#         else:
#             file_name = "media/data/" + files[0]
#             # Opening JSON file
#             f = open(file_name , encoding='utf-8')
#             # returns JSON object as  a dictionary
#             allData = json.load(f)

#             print("="*20, "allData already existing", "="*20)
#             # print(allData)
#     else:
#         list_file = os.listdir("media/data")
#         for file_name in list_file:
#             os.remove("media/data/" + str(file_name))

#         allData = subFunction(token, day, baseUrl)
#         print("="*20, "subFunction getting allData", "="*20)
#         # print(allData[5:])

#     return allData


# print(load_data(token = '9f49725e-2402-46fd-ab52-c03a6ba2c529',
#                 baseUrl = 'https://observe.ob7.ird.fr/observeweb/api/public'))


# def load_allData_file():
#     files = os.listdir("media/data")

#     file_name = "media/data/" + files[0]
#     # Opening JSON file
#     f = open(file_name,  encoding='utf-8')
#     # returns JSON object as  a dictionary
#     allData = json.load(f)

#     return allData

# def load_json_file(file_path):
#     try:
#         with open(file_path, 'r') as file:
#             data = json.load(file)
#             return data
#     except FileNotFoundError:
#         print(f"File '{file_path}' not found.")
#         return None
#     except json.JSONDecodeError as e:
#         print(f"Error decoding JSON file '{file_path}': {e}")
#         return None

# def serialize(obj): 
#     """ 
#     Serialize obj dans un format json de type date, int ou str.
#     """
#     if isinstance(obj, datetime.datetime):
#         return obj.isoformat()
#     if isinstance(obj, np.int64):
#         return int(obj)
#     return str(obj)
#     # raise TypeError("Type not serializable")

# def pretty_print(json_data, file="media/temporary_files/created_json_file.json", mode="a"):
#     """ Fonction qui affiche le fichier json avec les bonnes indentations un fichier json

#     Args:
#         json_data (json): Données json en entrée
#         file (str, optional): Nom de fichier json de sortie "created_json_file.json".
#         mode (str, optional): Defaults to "a" pour "append" - "w" pour "write"
#     """

#     json_formatted_str = json.dumps(
#         json_data, indent=2, default=serialize)
#     # print("¤"*20, "pretty print function" ,"¤"*20)
#     # print("pretty print type ::::", type(json_formatted_str), 'and before it was :::', type(json_data))
#     with open(file, mode) as outfile:
#         outfile.write(json_formatted_str)


########### Fonctions faisant appel au le web service ###########
# def getId_Data(token, url, moduleName, argment, route):
#     """
#     Permet de retourner un id en fonction du module et de la route envoyé
#     """
#     headers = {
#         "Content-Type": "application/json",
#         'authenticationToken': token
#     }

#     urls = url + route + moduleName + "?filters." + argment
#     rep = requests.get(urls, headers=headers)

#     # print(rep.url)

#     if rep.status_code == 200:
#         return json.loads(rep.text)["content"][0]["topiaId"]
#     else:
#         return json.loads(rep.text)["message"]


# def check_trip(token, content, url_base):
#     """
#     Permet de verifier si la marée a inserer existe déjà dans la base de donnée
#     """
#     start = content["startDate"].replace("T00:00:00.000Z", "")
#     end = content["endDate"].replace("T00:00:00.000Z", "")

#     vessel_id = content["vessel"].replace("#", "-")

#     # print(start, end, vessel_id)

#     id_ = ""
#     ms_ = True

#     try:
#         id_ = getId_Data(token, url=url_base, moduleName="Trip", route="/data/ps/common/",
#                         argment="startDate=" + start + "&filters.endDate=" + end + "&filters.vessel_id=" + vessel_id)
#     except:
#         ms_ = False

#     return id_, ms_


# def get_one_from_ws(token, url_base, route, topiaid):
#     """ Fonction qui interroge le web service (ws) pour récupérer toutes les données relatives à une route et un topiaid

#     Args:
#         token (str): token
#         url_base: chemin d'accès à la connexion ('https://observe.ob7.ird.fr/observeweb/api/public')
#         route: chemin d'accès plus précis (par ex : '/data/ll/common/Trip/')
#         topiaid: topiaid avec des '-' à la place des '#'

#     Returns:
#         file.json: informations relatives au topiaid fourni
#     """

#     headers = {
#         'authenticationToken': token, 
#     }

#     params = {
#         'config.recursive' : 'true', 
#     }

#     url = url_base + route + topiaid

#     response = requests.get(url, headers=headers, params = params, timeout=15)

#     if response.status_code == 200 :
#         # with open(file = "media/temporary_files/previoustrip.json", mode = "w") as outfile:
#         #     outfile.write(response.text)
#         return response.content

#     else:
#         return None


# def send_trip(token, data, url_base, route):
#     """ Fonction qui ajoute un trip (marée) dans la base

#     Args:
#         token (str): token
#         data (json): json file (trip) que l'on envoie dans la base
#         url_base (str): 'https://observe.ob7.ird.fr/observeweb/api/public' base de connexion à l'api
#         route (str):  '/data/ps/common/Trip' ou '/data/ll/common/Trip'
#     Returns:
#         le json inséré dans le temporary_files 
#         text message: logbook bien inséré, ou bien un json d'erreur
#     """

#     data_json = json.dumps(data)

#     headers = {
#         "Content-Type": "application/json",
#         'authenticationToken': token
#     }

#     url = url_base + route

#     print("Post")
#     pretty_print(data)

#     response = requests.post(url, data=data_json, headers=headers)

#     print(response.status_code, "\n")

#     if response.status_code == 200:
#         # return json.loads(res.text)
#         return ("Logbook inséré avec success", 1)
#     else:
#         with open(file = "media/temporary_files/error.json", mode = "w") as outfile:
#             outfile.write(response.text)
#         try:
#             return (errorFilter(response.text), 2)
#             # return json.loads(res.text), 2
#         except KeyError:
#             # Faire une fonction pour mieux traiter ce type d'erreur
#             # print("Message d'erreur: ", json.loads(res.text)["exception"]["result"]["nodes"]) # A faire
#             print("Message d'erreur: ", json.loads(response.text)) # A faire
#             return ("L'insertion de cet logbook n'est pas possible. Désolé veuillez essayer un autre", 3)


# def update_trip(token, data, url_base, topiaid):
#     """Fonction qui met à jour un trip dans la base de données, donc supprime le trip existant pour insérer le nouveau data_json sous le même topiaid

#     Args:
#         token (str): token
#         data (json): json file qu'on envoie dans la base
#         url_base (str): 'https://observe.ob7.ird.fr/observeweb/api/public' base de connexion à l'api
#         topiaid du trip que l'on veut update (l'ancienne version sera supprimée)
#     Returns:
#     """

#     data_json = json.dumps(data, default=serialize)

#     headers = {
#         "Content-Type": "application/json",
#         'authenticationToken': token,}

#     url = url_base + '/data/ll/common/Trip/' + topiaid

#     pretty_print(data)
#     response = requests.put(url, data=data_json, headers=headers, timeout=15)

#     print("Code resultat de la requete", response.status_code)

#     if response.status_code == 200:
#         return ("Logbook inséré avec success", 1)
#     else:
#         with open(file = "media/temporary_files/errorupdate.json", mode = "w") as outfile:
#             outfile.write(response.text)


# def trip_for_prog_vessel(token, url_base, route, vessel_id, programme_topiaid):
#     """
#     Pour un navire et un programme donnée, renvoie le topiaid du dernier trip saisi

#     Args:
#         token
#         url_base: 'https://observe.ob7.ird.fr/observeweb/api/public'
#         vessel_id: topiaid du navire (avec les '-')
#         programme_topiaid: topiaid du programme choisi (avec les '-')

#     Returns:
#         trip topiaid
#     """
#     # api_base = 'https://observe.ob7.ird.fr/observeweb/api/'
#     api_trip = '?authenticationToken='

#     api_vessel_filter = '&filters.vessel_id='
#     api_programme_filter = '&filters.logbookProgram_id='
#     api_ordeer_filter = '&orders.endDate=DESC'

#     api_trip_request = url_base + api_trip + token + api_vessel_filter + vessel_id + api_programme_filter + programme_topiaid + api_ordeer_filter
#     response = requests.get(api_trip_request, timeout=15)
#     return response.content

# ########### Fonctions de recherche dans allData --> à déplacer dans json_functions ###########

# Retourne ID d'un module en fonction des arguments donnés
def getId(allData, moduleName, argment, nbArg=False, domaine=None):
    """Fonction qui retourne l'ID d'un module en fonction des arguments donnés

    Args:
        allData (json): données de references
        moduleName (str): le module de la base de donnée
        argment (str): les arguments de la requete sur le module
        domaine (str): "seine" ou "longline" dans le cas ou nous voulons recuperer les id de VesselActivity
        nbArg (bool): permet de signifier le nombre d'argument dont on aura besoin pour trouver l'ID
               par defaut quand c'est False nous avons 1 argument en paramentre
               si c'est True, nous avons 2 arguments en parametre

    Returns:
         topiad (str)
    """
    message = ""
    Id = ""
    dataKey = [k for (k, v) in allData.items()]

    if moduleName in dataKey:
        if domaine != None:
            tempDic = allData[moduleName][domaine]
        else:
            tempDic = allData[moduleName]

        if nbArg:
            # 2 arguments
            argTab = argment.split("&filters.")
            argments = [argTab[0].split("="), argTab[1].split("=")]
            for val in tempDic:
                if (val[argments[0][0]] == argments[0][1]) and (val[argments[1][0]] == argments[1][1]):
                    Id = val['topiaId']
        else:
            # 1 argument
            argments = argment.split("=")
            for val in tempDic:
                if val[argments[0]] == argments[1]:
                    Id = val['topiaId']

        if Id == "":
            # message = "Aucun topiad"
            Id = None
    else:
        # message = "Le module: "+ module + " n'existe pas"
        Id = None

    return Id


def getSome(allData, moduleName, argment):
    """Permet de retouner un dictionnaire de donnée du module dans une liste (tableau)

    Args:
        allData (json): données de references
        moduleName (str): le module de la base de donnée
        argment (str): les arguments de la requete sur le module

    Returns:
        (list)
    """
    tempDic = {}
    dico = {}
    dataKey = [k for (k, v) in allData.items()]

    if moduleName in dataKey:
        tempDic = allData[moduleName]
        # print(tempDic)
        argments = argment.split("=")
        for val in tempDic:
            if val[argments[0]].lower() == argments[1].lower():
                dico = val

    return [dico]


def getAll(allData, moduleName, type_data="dictionnaire"):
    """Permet de retourner un dictionnaire ou un tableau

    Args:
        allData (json): données de references
        moduleName (str): le module de la base de donnée
        type_data (str): "dictionnaire" ou "tableau"

    Returns:
        tab (list)
        dico (dict)
    """
    if type_data == "tableau":
        tab = []
        for val in allData[moduleName]:
            tab.append((val["topiaId"], val["label1"]))

        return tab
    else:
        dico = {}
        for val in allData[moduleName]:
            dico[val["code"]] = val["topiaId"]

        return dico


# Traitement du logbook
def traiLogbook(logB):
    """Fonction qui permet d'extraire les informations d'un logbook SENNE dans ces variables
        --> info_bat : contient les informations sur le navire, le port (D/A) et heure (D/A)
        --> observ : contient les informations sur le capitaine et le mar_homeid
        --> df_data : contient les données du logbook

    Args:
        logB (str): fichier logbook à traiter

    Returns:
         info_bat (json), df_data (dataframe), observ (json), (str)
    """

    # Chargement du fichier
    # wb = Workbook()

    try:
        wb = op.load_workbook(logB)
    except Exception as e:
        return '', '', '', "Error : Fichier illisible" + str(e)

    if "xlsm" in logB.split('.') or "XLSM" in logB.split('.'):
        # Recuperer le non du bateau et autres information utils
        #   st.text(wb.get_sheet_names())
        maree_log = wb["1.Marée"]

        # Recuperer la feuille active
        act_sheet = wb["2.Logbook"]
    else:
        # Recuperer le non du bateau et autres information utils
        maree_log = wb["Marée"]
        # st.text(wb.get_sheet_names())

        # Recuperer la feuille active
        act_sheet = wb["Logbook"]

    info_bat = {
        "Navire": maree_log["F"][1].value,
        "Depart_Port": maree_log["F"][12].value,
        "Depart_Date": str(maree_log["F"][13].value).split(" ")[0],
        "Depart_heure": str(maree_log["F"][14].value).split(" ")[0],
        "Arrivee_Port": maree_log["F"][17].value,
        "Arrivee_Date": str(maree_log["F"][18].value).split(" ")[0],
        "Arrivee_Loch": maree_log["F"][20].value,
    }

    observ = {
        "captain": maree_log["D"][9].value,
        "mar_homeId": maree_log["D"][10].value,
    }

    # Variable pour recuperer les donnée dans le logbook
    data = []
    obj = []

    # Recuperation des lignes qui nous interesses à partir de la ligne 33 dans le fichier
    i = 1
    for row in act_sheet.rows:
        if i >= 33:
            for index in range(len(row)):
                obj.append(row[index].value)
            data.append(obj)
            obj = []
        i = i + 1

    # Transformer le tableau "data" en dataFrame pour faciliter la manipulation des données
    data = pd.DataFrame(np.array(data))

    # Titrer le tableau
    data = data.rename(
        columns={0: "date", 1: "heure", 2: "lat1", 3: "lat2", 4: "lat3", 5: "long1", 6: "long2", 7: "long3", 8: "zee",
                 9: "temp_mer", 10: "vent_dir", 11: "vent_vit", 12: "calee_porta", 13: "calee_nul", 14: "calee_type",
                 15: "cap_alb_yft_p10_tail", 16: "cap_alb_yft_p10_cap", 17: "cap_alb_yft_m10_tail",
                 18: "cap_alb_yft_m10_cap", 19: "cap_lst_skj_tail", 20: "cap_lst_skj_cap", 21: "cap_pat_bet_p10_tail",
                 22: "cap_pat_bet_p10_cap", 23: "cap_pat_bet_m10_tail", 24: "cap_pat_bet_m10_cap",
                 25: "cap_ger_alb_tail", 26: "cap_ger_alb_cap", 27: "cap_aut_esp_oth_esp", 28: "cap_aut_esp_oth_tail",
                 29: "cap_aut_esp_oth_cap", 30: "cap_rej_dsc_esp", 31: "cap_rej_dsc_tail", 32: "cap_rej_dsc_cap",
                 33: "asso_bc_libre", 34: "asso_objet", 35: "asso_balise", 36: "asso_baliseur", 37: "asso_requin",
                 38: "asso_baleine", 39: "asso_oiseaux", 40: "obj_flot_act_sur_obj", 41: "obj_flot_typ_obj",
                 42: "obj_flot_typ_dcp_deriv", 43: "obj_flot_risq_mail_en_surf", 44: "obj_flot_risq_mail_sou_surf",
                 45: "bouee_inst_act_bou", 46: "bouee_inst_bou_prst_typ", 47: "bouee_inst_bou_prst_id",
                 48: "bouee_inst_bou_deplo_typ", 49: "bouee_inst_bou_deplo_id", 50: "comment"})

    #####  Traitement pour supprimer les lignes qui n'ont pas de donnée dans le datFrame 'data'

    # Suppression des lignes identiques c.a.d les doublons
    df_data = data.drop_duplicates(keep=False)

    df_data.loc[:, "date"] = df_data.date.fillna(method="ffill")

    # Si nous avons des ligne contenant des valeurs NaT; les ignorer et garder la bonne données
    if df_data.date.isnull().sum() > 0:
        df_data = df_data[~df_data["date"].isna()]
        df_data.reset_index(drop=True,
                            inplace=True)  #  réinitialiser l'index à son format par défaut (c'est-à-dire un RangeIndex de 0 à la longueur du cadre de données moins 1)

    df_data = df_data.loc[:, :"comment"]

    if info_bat['Depart_Date'] == 'None':
        info_bat['Depart_Date'] = str(df_data['date'].iloc[0]).split(' ')[0]

    if info_bat['Arrivee_Date'] == 'None':
        info_bat['Arrivee_Date'] = str(df_data['date'].iloc[-1]).split(' ')[0]

    return info_bat, df_data, observ, ''


# Traitement du logbook orth v23
def traiLogbook_v23(logB):
    """Fonction qui permet d'extraire les informations d'un logbook SENNE dans ces variables
        --> info_bat : contient les informations sur le navire, le port (D/A) et heure (D/A)
                      et contient les informations sur le capitaine et le mar_homeid
        --> df_data : contient les données du logbook

    Args:
        logB (str): fichier logbook à traiter

    Returns:
         info_bat (json), df_data (dataframe), (str)
    """

    try:
        wb = op.load_workbook(logB)
    except Exception as e:
        print("Error :", e)

    maree_log = wb[wb.sheetnames[3]]
    act_sheet = wb[wb.sheetnames[4]]

    info_bat = {
        "Navire": maree_log["C"][3].value,
        "Depart_Port": maree_log["C"][17].value,
        "Depart_Date": str(maree_log["C"][18].value).split(" ")[0],
        "Depart_heure": str(maree_log["C"][19].value).split(" ")[0],
        # "Depart_heure": str(maree_log["C"][19].value),
        "Arrivee_Port": maree_log["C"][22].value,
        "Arrivee_Date": str(maree_log["C"][23].value).split(" ")[0],
        "Arrivee_Loch": maree_log["C"][25].value,
        "captain": str(maree_log["C"][11].value) + " " + str(maree_log["C"][12].value),
        "mar_homeId": str(maree_log["C"][14].value) + "" + str(maree_log["C"][15].value),
    }

    # Variable pour recuperer les donnée dans le logbook
    data = []
    obj = []

    # Recuperation des lignes qui nous interesses à partir de la ligne 33 dans le fichier

    i = 1
    for row in act_sheet.rows:
        if i >= 7:
            for index in range(len(row)):
                obj.append(row[index].value)
            data.append(obj)
            obj = []
        i = i + 1

    # Transformer le tableau "data" en dataFrame pour faciliter la manipulation des données
    data = pd.DataFrame(np.array(data))

    try:
        data = data.drop(data.columns[[0, data.shape[1]-2]], axis=1)  # supprimer la 1ere et derniere colonne
        data.columns = range(data.columns.size)  # reninitialiser les colonnes à partir de 0

        # Titrer le tableau
        data = data.rename(
            columns={0: "type_declaration", 1: "date", 2: "heure",
                     3: "lat1", 4: "lat2", 5: "lat3", 6: "long1",
                     7: "long2", 8: "long3", 9: "port", 10: "zee",
                     11: "temp_mer", 12: "vent_dir", 13: "vent_vit",
                     14: "calee_type", 15: "espece", 16: "categ_poids",
                     17: "quant_conser_tonne", 18: "quant_conser_nb",
                     19: "quant_reje_tonne", 20: "quant_reje_nb",
                     21: "obj_flot_act_sur_obj", 22: "obj_flot_typ_obj",
                     23: "obj_flot_typ_dcp_deriv", 24: "obj_type_composant",
                     25: "obj_nombre", 26: "obj_hauteur", 27: "obj_longueur",
                     28: "obj_largeur", 29: "obj_profondeur", 30: "obj_mailles",
                     31: "obj_plastique", 32: "obj_metal", 33: "obj_bio",
                     34: "bouee_inst_act_bou", 35: "bouee_posit_connue",
                     36: "bouee_nav_proprietaire", 37: "bouee_modele", 38: "bouee_numero", 39: "commentaire"})

        #####  Traitement pour supprimer les lignes qui n'ont pas de donnée dans le datFrame 'data'

        # Suppression des lignes identiques c.a.d les doublons
        df_data = data.drop_duplicates(keep=False)

        # pd.options.mode.copy_on_write = True
        df_data.date = df_data.date.fillna(method="ffill")

        # Si nous avons des ligne contenant des valeurs NaT; les ignorer et garder la bonne données
        if df_data.date.isnull().sum() > 0:
            df_data = df_data[~df_data["date"].isna()]
            df_data.reset_index(drop=True,
                                inplace=True)  #  réinitialiser l'index à son format par défaut (c'est-à-dire un RangeIndex de 0 à la longueur du cadre de données moins 1)
        df_data = df_data.loc[:, :"commentaire"]

        return info_bat, df_data, ""

    except:
        return None, None, "Mauvais format de logbook"


def read_data(file, type_doc="v21"):
    """Fonction qui permet de faire appel à la fonction de traitement du logbook

    Args:
        file (str): fichier logbook à traiter
        type_doc (str): logbook "v21" ou "v23"

    Returns:
         info_bat (json), data_bat (dataframe), obs (json), message (str)
    """

    if type_doc == "v21":
        info_bat, data_bat, obs, message = traiLogbook(file)
        return info_bat, data_bat, obs, message
    if type_doc == "v23":
        info_bat, data_bat, message = traiLogbook_v23(file)
        return info_bat, data_bat, message


def get_lat_long(allData, harbour):
    """Fonction qui permet de retourner les coordonnées de longitude et de latitude du port de depart ou soit d'arrivé

    Args:
        allData (json): données de references
        harbour (str): nom du port

    Returns:
         (float), (float)
    """
    print(harbour)
    if harbour != None:
        harbour = str(harbour).strip()
        for val in allData["Harbour"]:
            # print("Label 2: ",val["label2"], " ==> Recherche: ", harbour)

            if (harbour.lower() in val["label1"].lower()) or (harbour.lower() in val["label2"].lower()) or (
                    harbour.lower() in val["label3"].lower()):
                return float(val["latitude"]), float(val["longitude"])

        # return "Le port de départ << "+ harbour + " >> n'a pas été trouvé dans le service."
        return None, None
    else:
        # return "Le port de départ << "+ harbour + " >> n'a pas été trouvé dans le service."
        return None, None


def lat_long(lat1, lat2, lat3, long1, long2, long3):
    """Fonction qui permet de calculer la longitude et la latitude pour l'inserer facilement dans la BD

    Args:
        lat1, lat2, lat3, long1, long2, long3 (str): position geographique

    Returns:
         (float), (float), (bool)
    """
    try:
        lat_1 = int(float(str(lat1).replace("°", "")))
        lat_2 = int(float(str(lat2).replace("'", "")))
        long_1 = int(float(str(long1).replace("°", "")))
        long_2 = int(float(str(long2).replace("'", "")))

        # Latitude
        var1 = (lat_2 / 60)
        var1 = str(var1)
        varAr1 = var1.split(".")

        # Longitude
        var2 = (long_2 / 60)
        var2 = str(var2)
        varAr2 = var2.split(".")

        if lat_2 >= 60:
            lat_ = varAr1[0] + varAr1[1][:4]
            add = int(lat_1) + float(lat_)
            res_long = str(add)
        else:
            lat_ = varAr1[1][:4]
            res_lat = str(lat_1) + "." + lat_

        if long_2 >= 60:
            long_ = varAr2[0] + varAr2[1][:4]
            add = int(long_1) + float(long_)
            res_long = str(add)
        else:
            long_ = varAr2[1][:4]
            res_long = str(long_1) + "." + long_

        # print(res_long)
        if (lat3 != None) and (long3 != None):
            if lat3.lower() == "s":
                res_lat = float(res_lat) * (-1)
            else:
                res_lat = float(res_lat)

            if long3.lower() == "w":
                res_long = float(res_long) * (-1)
            else:
                res_long = float(res_long)

            return res_lat, res_long, False
        else:
            return None, None, True
    except ValueError:
        return None, None, False

def get_wind_id_interval(allData, moduleName, windSpeed):
    """Fonction qui permet de retourner le topiaId de l'interval de vitesse du vent

    Args:
        allData (json): données de references
        moduleName (str): le module de la base de donnée
        windSpeed (str): vitesse du vent

    Returns:
         id (str)
    """

    tab = []
    for val in allData[moduleName]:
        if val['code'] == '0':
            pass
        else:
            try:
                if (int(val['minSpeed']) <= int(windSpeed) <= int(val['maxSpeed'])): return val['topiaId']
            except:
                return None
    return None

def weightCategory(allData, chaine, specise):
    """Fonction qui permet de retourner le topiaId de la categorie d'especes trouvées

    Args:
        allData (json): données de references
        chaine (str): chaine contenant la categorie
        specise (str): type d'especes

    Returns:
         id (str)
    """

    inconnu = [val['topiaId'] for val in allData["WeightCategory"] if 'inconnu' in val['label2'].lower()]
    if chaine != None:
        if ">" in chaine:
            chaine  = chaine.replace(">", "plus de")
        elif "<" in chaine:
            chaine  = chaine.replace("<", "moins de")
        elif "-" in chaine:
            chaine  = chaine.replace("-", "à")


        for val in allData["WeightCategory"]:
            if ((specise.lower() in val['label2'].lower()) and (chaine.lower() in val['label2'].lower())):
                # print("Chaine: ", specise, chaine, " Id & lab: ",val['topiaId'], val['label2'])
                return val['topiaId']
        return inconnu[0]
    else:
        return inconnu[0]

def fpaZone_id(chaine, tableau, allData):
    """Fonction qui permet de retourner le topiaId et un commentaire lorsqu'on ne retrouve pas la zone fpa passée en paramettre

    Args:
        allData (json): données de references
        tableau (list): liste de données fpa zone
        chaine (str): zone à rechercher

    Returns:
         id (str), '' (str) ou chaine (str)
    """

    status = False
    for val in tableau:
        # print("FPA : ", chaine)
        if (chaine != None) and (chaine.lower() in val[1].lower()):
            return val[0], ""
        status = True
    if status:
        return getId(allData, "FpaZone", argment="code=XXX*"), chaine


def transmittingBType(chaine, dico):
    """Fonction qui permet de retourner le topiaId du type de balise et un commentaire lorsqu'on ne retrouve pas la balise passée en paramettre

    Args:
        dico (json): dictionnaire de données pour les balises
        chaine (str): balise à rechercher

    Returns:
         id (str), '' (str) ou commentaire (str)
    """
    if ("m3i+" in str(chaine).lower()): return dico['26'], ""
    if ("m3igo" in str(chaine).lower()): return dico['29'], ""
    if ("m3i" in str(chaine).lower()): return dico['25'], ""
    if ("m4i+" in str(chaine).lower()): return dico['28'], ""
    if ("m4i" in str(chaine).lower()): return dico['27'], ""
    if ("thalos mod" in str(chaine).lower()): return dico['90'], ""
    if ("ortbit+" in str(chaine).lower()): return dico['92'], ""
    if ("orbit+" in str(chaine).lower()): return dico['92'], ""
    if ("ortbit" in str(chaine).lower()): return dico['91'], ""
    if ("orbit" in str(chaine).lower()): return dico['91'], ""
    if ("marine instru" in str(chaine).lower()): return dico['20'], ""
    if ("SLX+" in str(chaine).lower()): return dico['47'], ""
    if ("slx+" in str(chaine).lower()): return dico['47'], ""
    if ("isd+" in str(chaine).lower()): return dico['46'], ""
    if ("autre satl" in str(chaine).lower()): return dico['40'], ""
    if ("satlink mod" in str(chaine).lower()): return dico['40'], ""

    if ("marque et mod" in str(chaine).lower()): return dico['98'], ""
    if ("pas de bou" in str(chaine).lower()): return dico['999'], ""

    if chaine != "":
        # Getsome  Zunibal
        # https://demo.ultreia.io/observe-9.1.6/doc/api/public/
        # print(dico)
        # Si n'est pas present dans le Getsome
        return dico['99'], chaine + ": Balise // Type à préciser"
    else:
        # auccun type de balise dans le logbook
        return dico['999'], "Aucun type de balise indiqué dans le logbook"


def floatingObjectPart(chaine, data, dico, index, perte_act=False):
    """Fonction qui permet de retourner un ou deux topiaId en fonction du type d'objet flottant

    Args:
        data (dataFrame): les données du logbook en cours d'insertion
        index (str): nom de la colonne du dataFrame faisant reference aux objets flottants
        dico (json): dictionnaire de données pour les objets flottants
        perte_act (bool): gestion des pertes
        chaine (str): objet flottant à rechercher

    Returns:
         id (str), id (str)
    """

    # Types objets flottants
    if index == 'obj_flot_typ_obj':
        if ("dcp ancré" in str(chaine).lower()): return dico['1-2']
        if (("epave artificielle liée à la pêche" in str(chaine).lower()) or ("débris artificiel issu de la pêche" in str(chaine).lower())): return dico['2-2-4']
        if (("epave artificielle liée à d'autres activités humaines" in str(chaine).lower()) or ("débris artificiel non issu de la pêche" in str(chaine).lower())): return dico['2-2-5']
        if ("d'origine animale" in str(chaine).lower()): return dico['2-1-2']
        if ("d'origine végétale" in str(chaine).lower()): return dico['2-1-1']


        if ("dcp dérivant" in str(chaine).lower()):
            # Types de DCP
            if (("dcp français émergé bambou" in str(data['obj_flot_typ_dcp_deriv']).lower()) or ("radeau émergé bambou" in str(data['obj_flot_typ_dcp_deriv']).lower())): return dico['1-1-1-1-1']
            if ("radeau immergé métal" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['1-1-1-1-5']
            if ("dcp français émergé métal" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['1-1-1-1-2']
            if ("dcp français émergé bambou-métal" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico[
                '1-1-1-1-1'], dico['1-1-1-1-2']
            if ("dcp français furtif" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['4-9'], dico['1-1-1']
            if ("dcp français cage" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['1-1-1']
            if ("dcp espagnol émergé bambou" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['1-1-1-1-1']
            if ("dcp espagnol émergé métal" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['1-1-1-1-2']
            if ("traine" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['1-1-2-6']
            if ("dcp espagnol émergé bambou-métal" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico[
                '1-1-1-1-1'], dico['1-1-1-1-2']
            if ("dcp espagnol émergé plastique" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico[
                '1-1-1-1-2']
            if ("dcp espagnol furtif" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['4-9'], dico['1-1-1']
            if ("dcp espagnol cage" in str(data['obj_flot_typ_dcp_deriv']).lower()) or ("radeau" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['1-1-1']
            if ("dcp coréen" in str(data['obj_flot_typ_dcp_deriv']).lower()) : return dico['1-1']

            if (("dcp cage" in str(data['obj_flot_typ_dcp_deriv']).lower()) or \
                ("dcp avd" in str(data['obj_flot_typ_dcp_deriv']).lower()) or \
                ("autre dcp dérivant" in str(data['obj_flot_typ_dcp_deriv']).lower()) or \
                ("autre objet" in str(data['obj_flot_typ_dcp_deriv']).lower()) or \
                (data['obj_flot_typ_dcp_deriv'].isna() | (data['obj_flot_typ_dcp_deriv'] == "")).any()):

                return dico['1-1']

            if ("cage" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['1-1-2-5']
            if ("tas de bout" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['2-2-4-1']
            if ("tas de paille" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['2-1-1-1']
            if ("bille de bois" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['2-1-1-3']
            if ("defense" in str(data['obj_flot_typ_dcp_deriv']).lower()): return dico['2-2']

    # Mailles
    if index == 'obj_mailles':
        if ("pas de mailles" in str(chaine).lower()): return dico['1-1-1-2-1-3']
        if ("< 7 cm" in str(chaine).lower()): return dico['1-1-1-2-1-1']
        if ("> 7 cm" in str(chaine).lower()): return dico['1-1-1-2-1-2']
        if ("mailles de taille inconnue" in str(chaine).lower()): return dico['1-1-1-2-1-6']
        if ("non observable" in str(chaine).lower()): return dico['1-1-1-2-1-5']

    # Risque de maillage en surface
    if index == 'obj_flot_risq_mail_en_surf':
        if ("pas de mailles" in str(chaine).lower()): return dico['1-1-1-2-3']
        if ("< 6, 5 cm" in str(chaine).lower()): return dico['1-1-1-2-1']
        if ("> 6,5 cm" in str(chaine).lower()): return dico['1-1-1-2-2']
        if ("mailles de taille inconnue" in str(chaine).lower()): return dico['1-1-1-2-5']
        if ("non observable" in str(chaine).lower()): return dico['1-1-1-2-5']

    # Risque de maillage sous la surface
    if index == 'obj_flot_risq_mail_sou_surf':
        if ("pas de mailles" in str(chaine).lower()): return dico['1-1-1-3-3'], dico['1-1-2-3']
        if ("< 6, 5 cm" in str(chaine).lower()): return dico['1-1-1-3-1'], dico['1-1-2-4-2']
        if ("> 6,5 cm" in str(chaine).lower()): return dico['1-1-1-3-2'], dico['1-1-2-4-3']
        if ("mailles de taille inconnue" in str(chaine).lower()): return dico['1-1-1-3-5'], dico['1-1-2-2']
        if ("non observable" in str(chaine).lower()): return dico['1-1-1-3-5'], dico['1-1-2-2']

    # Autre DFAD
    if perte_act:
        return dico['1-1']


def cap_obs_sea(allData, ob):
    """Fonction qui permet d'appliquer des traitement sur le nom et prenoms du capitaine pour retrouver son topiaId

    Args:
        allData (json): données de references
        ob (str): les informations sur le capitaine et le mar_homeid

    Returns:
         id (str)
    """
    cap = []
    logbookDataEntryOperator = []

    for val in allData["Person"]:
        if val["captain"] == True:
            cap.append((val["topiaId"], val["firstName"].lower(), val["lastName"].lower()))
        elif val["dataEntryOperator"] == True:
            logbookDataEntryOperator.append((val["topiaId"], val["firstName"].lower(), val["lastName"].lower()))

    def sou_fonc(arra):
        print(ob['captain'])
        trouv_id = None
        inconnu_id = [val[0] for val in arra if (("[inconnu]" == val[1]) and ("[inconnu]" == val[2]))]

        if ob['captain'] != None:
            if len(ob['captain'].split(" ")) > 2:
                nom = ob['captain'].split(" ")[0]
                lisPren = ob['captain'].split(" ")[1:]
                prenoms = " ".join(lisPren)
            else:
                status = 0
                try:
                    # lorsqu'on a le nom et prenoms; ex: paul kenji
                    nom, prenoms = ob['captain'].split(" ")

                except ValueError:
                    # lorsque nous avons une seule informations saisie soit le nom ou le prenom; ex: kenji
                    nom_prenoms = ob['captain']

            if status == 1:
                trouv_id = [val[0] for val in arra if ((nom.lower() in str(val)) and (prenoms.lower() in str(val)))]
                if trouv_id != "": return trouv_id[0]

            elif status == 2:
                # Lorsqu'on a le nom seulement dans la base de donnée
                trouv_id = [val[0] for val in arra if ((nom_prenoms.lower() in str(val)) and ("[inconnu]" == val[2]))]
                if trouv_id != "": return trouv_id[0]

                # Lorsqu'on a le prenom seulement dans la base de donnée
                trouv_id = [val[0] for val in arra if (("[inconnu]" == val[1]) and (nom_prenoms.lower() in str(val)))]
                if trouv_id != "": return trouv_id[0]
            else:
                return inconnu_id[0]
        else:
            return inconnu_id[0]

    if cap == logbookDataEntryOperator:
        id_cap = sou_fonc(cap)
        return id_cap, id_cap
    else:
        id_cap = sou_fonc(cap)
        # id_op = sou_fonc(logbookDataEntryOperator) # Pas utilisé pour l'instant

        return id_cap, id_cap


class TransmitException(Exception):
    """Classe qui permet de gerer les exceptions sur les activités sur objet

    Args:

    Returns:

    """

    def __init__(self, message):
        self.message = message
        super().__init__(self.message)


def obj_deja_deploy(data, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner, allData, operation):
    """Fonction qui permet de construire les operations sur les objets et retourne l'ensemble des operations sur balise pour une route

    Args:
        dico_trams (json): dictionnaire de données pour les balises
        dico_trams_oper (json): dictionnaire de données pour les operations sur balises
        dico_trams_owner (json): dictionnaire de données pour les  sur balises
        allData (json): données de references
        js_Transmitts (json): structure json pour un enregistrement sur les balises
        data (dataFrame): les données du logbook en cours d'insertion
        operation (str) : permet de connaitre le type d'operation (visite, transfert, mise à l'eau...)

    Returns:
         tab2_Transmitt (liste) : liste de structure de données Json
    """
    tab2_Transmitt = []

    def sub_func(data, tab_dcp_type_and_id, code_trams_oper, code_trams_owner, dico_trams_oper, dico_trams,
                 dico_trams_owner, js_Transmitts):
        js_Transmitts['transmittingBuoyOperation'] = dico_trams_oper[code_trams_oper]
        js_Transmitts['transmittingBuoyType'], comment = transmittingBType(data[tab_dcp_type_and_id[0]], dico_trams)
        if comment != "":
            js_Transmitts['comment'] = comment

        if data[tab_dcp_type_and_id[1]] != "":
            if data[tab_dcp_type_and_id[1]] != None:
                code = int(data[tab_dcp_type_and_id[1]])
                if code == 0:
                    js_Transmitts['code'] = None
                else:
                    js_Transmitts['code'] = str(code)
            else:
                code = str(data[tab_dcp_type_and_id[1]])
                if (code == "None") or (code == "0"):
                    js_Transmitts['code'] = None
                else:
                    js_Transmitts['code'] = str(data[tab_dcp_type_and_id[1]])
        else:
            js_Transmitts['code'] = None

        js_Transmitts['transmittingBuoyOwnership'] = dico_trams_owner[code_trams_owner]

        return js_Transmitts

    def fun_dcp_activ(data, Basetab_dcp_type_and_id, Othtab_dcp_type_and_id, code_trams_oper, code_trams_owner,
                      dico_trams_oper, dico_trams, dico_trams_owner, allData, js_Transmitts, makeException=True):

        if (data[Basetab_dcp_type_and_id[0]] != None) and (
                "pas de" not in str(data[Basetab_dcp_type_and_id[0]]).lower()):
            js_transmi = sub_func(data, Basetab_dcp_type_and_id, code_trams_oper, code_trams_owner, dico_trams_oper,
                                  dico_trams, dico_trams_owner, js_Transmitts)

            if (data[Othtab_dcp_type_and_id[0]] != None) and (
                    "pas de" not in str(data[Othtab_dcp_type_and_id[0]]).lower()) and (makeException == True):
                # exception
                message = "Le " + str(data["date"]) + " à " + str(data[
                                                                      "heure"]) + " ===> Attention: Information non conforme sur le ou les FOB de l'activité: - '" + operation
                raise TransmitException(message)

            return js_transmi

        elif (data[Othtab_dcp_type_and_id[0]] != None) and (
                "pas de" not in str(data[Othtab_dcp_type_and_id[0]]).lower()):
            return sub_func(data, Othtab_dcp_type_and_id, code_trams_oper, code_trams_owner, dico_trams_oper,
                            dico_trams, dico_trams_owner, js_Transmitts)

    js_Transmitts = js_Transmitt()  # intialisatiion des parametres par defaut
    if ("mise" in str(data['bouee_inst_act_bou']).lower()):
        js_transmi = fun_dcp_activ(data, ['bouee_inst_bou_deplo_typ', 'bouee_inst_bou_deplo_id'],
                                   ['bouee_inst_bou_prst_typ', 'bouee_inst_bou_prst_id'], '3', '3', dico_trams_oper,
                                   dico_trams, dico_trams_owner, allData, js_Transmitts)
        tab2_Transmitt.append(js_transmi)

    js_Transmitts = js_Transmitt()  # intialisatiion des parametres par defaut
    if ("visite" in str(data['bouee_inst_act_bou']).lower()):
        js_transmi = fun_dcp_activ(data, ['bouee_inst_bou_prst_typ', 'bouee_inst_bou_prst_id'],
                                   ['bouee_inst_bou_deplo_typ', 'bouee_inst_bou_deplo_id'], '1', '3', dico_trams_oper,
                                   dico_trams, dico_trams_owner, allData, js_Transmitts)
        tab2_Transmitt.append(js_transmi)

    js_Transmitts = js_Transmitt()  # intialisatiion des parametres par defaut
    if (("retrait" in str(data['bouee_inst_act_bou']).lower()) or ("perte" in str(data['bouee_inst_act_bou']).lower())):
        js_transmi = fun_dcp_activ(data, ['bouee_inst_bou_prst_typ', 'bouee_inst_bou_prst_id'],
                                   ['bouee_inst_bou_deplo_typ', 'bouee_inst_bou_deplo_id'], '2', '3', dico_trams_oper,
                                   dico_trams, dico_trams_owner, allData, js_Transmitts)
        tab2_Transmitt.append(js_transmi)

        #############################
        # print("perte ou retrait ",data["date"])

    js_Transmitts = js_Transmitt()  # intialisatiion des parametres par defaut
    if ("fin" in str(data['bouee_inst_act_bou']).lower()):
        js_transmi = fun_dcp_activ(data, ['bouee_inst_bou_prst_typ', 'bouee_inst_bou_prst_id'],
                                   ['bouee_inst_bou_deplo_typ', 'bouee_inst_bou_deplo_id'], '5', '3', dico_trams_oper,
                                   dico_trams, dico_trams_owner, allData, js_Transmitts)
        tab2_Transmitt.append(js_transmi)

    js_Transmitts = js_Transmitt()  # intialisatiion des parametres par defaut
    if ("transfert" in str(data['bouee_inst_act_bou']).lower()):
        js_transmi = fun_dcp_activ(data, ['bouee_inst_bou_prst_typ', 'bouee_inst_bou_prst_id'],
                                   ['bouee_inst_bou_deplo_typ', 'bouee_inst_bou_deplo_id'], '2', '3', dico_trams_oper,
                                   dico_trams, dico_trams_owner, allData, js_Transmitts, makeException=False, )
        tab2_Transmitt.append(js_transmi)

        js_Transmitts = js_Transmitt()  # intialisatiion des parametres par defaut
        js_transmi = fun_dcp_activ(data, ['bouee_inst_bou_deplo_typ', 'bouee_inst_bou_deplo_id'],
                                   ['bouee_inst_bou_prst_typ', 'bouee_inst_bou_prst_id'], '3', '3', dico_trams_oper,
                                   dico_trams, dico_trams_owner, allData, js_Transmitts, makeException=False, )
        tab2_Transmitt.append(js_transmi)

    return tab2_Transmitt

def obj_deja_deploy_v23(data_ob, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner, allData, operation):
    """Fonction qui permet de construire les operations sur les objets et retourne l'ensemble des operations sur balise pour une route

    Args:
        dico_trams (json): dictionnaire de données pour les balises
        dico_trams_oper (json): dictionnaire de données pour les operations sur balises
        dico_trams_owner (json): dictionnaire de données pour les  sur balises
        allData (json): données de references
        js_Transmitts (json): structure json pour un enregistrement sur les balises
        data_ob (dataFrame): les données du logbook sur operation sur objet
        operation (str) : permet de connaitre le type d'operation (visite, transfert, mise à l'eau...)

    Returns:
         tab2_Transmitt (liste) : liste de structure de données Json
    """
    tab2_Transmitt = []

    def sub_func(data, tab_dcp_type_and_id, code_trams_oper, code_trams_owner, dico_trams_oper, dico_trams,
                 dico_trams_owner, allData, js_Transmitts):
        js_Transmitts['transmittingBuoyOperation'] = dico_trams_oper[code_trams_oper]
        js_Transmitts['transmittingBuoyType'], comment = transmittingBType(data[tab_dcp_type_and_id[0]], dico_trams)
        if comment != "":
            js_Transmitts['comment'] = comment

        if data[tab_dcp_type_and_id[1]] != "":
            if data[tab_dcp_type_and_id[1]] != None:
                code = int(data[tab_dcp_type_and_id[1]])
                if code == 0:
                    js_Transmitts['code'] = None
                else:
                    js_Transmitts['code'] = str(code)
            else:
                code = str(data[tab_dcp_type_and_id[1]])
                if (code == "None") or (code == "0") :
                    js_Transmitts['code'] = None
                else:
                    js_Transmitts['code'] = str(data[tab_dcp_type_and_id[1]])
        else:
            js_Transmitts['code'] = None

        js_Transmitts['transmittingBuoyOwnership'] = dico_trams_owner[code_trams_owner]

        return js_Transmitts

    def fun_dcp_activ(data, Basetab_dcp_type_and_id, Othtab_dcp_type_and_id, code_trams_oper, code_trams_owner,
                      dico_trams_oper, dico_trams, dico_trams_owner, allData, js_Transmitts, makeException=True):


        if (data[Othtab_dcp_type_and_id[0]] != None) and (data[Basetab_dcp_type_and_id[0]] == data[Othtab_dcp_type_and_id[0]]) and (
                "pas de" not in str(data[Othtab_dcp_type_and_id[0]]).lower()):              #############  Orth v23
            return sub_func(data, Othtab_dcp_type_and_id, code_trams_oper, code_trams_owner, dico_trams_oper,
                            dico_trams, dico_trams_owner, allData, js_Transmitts)

        elif (data[Basetab_dcp_type_and_id[0]] != None) and (
                "pas de" not in str(data[Basetab_dcp_type_and_id[0]]).lower()):
            js_transmi = sub_func(data, Basetab_dcp_type_and_id, code_trams_oper, code_trams_owner, dico_trams_oper,
                                  dico_trams, dico_trams_owner, allData, js_Transmitts)

            if (data[Othtab_dcp_type_and_id[0]] != None) and (
                    "pas de" not in str(data[Othtab_dcp_type_and_id[0]]).lower()) and (makeException == True):
                # exception
                message = "Le " + str(data["date"]) + " à " + str(data[
                                                                      "heure"]) + " ===> Attention: Information non conforme sur le ou les FOB de l'activité: - '" + operation
                raise TransmitException(message)

            return js_transmi

        elif (data[Othtab_dcp_type_and_id[0]] != None) and (
                "pas de" not in str(data[Othtab_dcp_type_and_id[0]]).lower()):
            return sub_func(data, Othtab_dcp_type_and_id, code_trams_oper, code_trams_owner, dico_trams_oper,
                            dico_trams, dico_trams_owner, allData, js_Transmitts)

    for _, data in data_ob.iterrows():

        js_Transmitts = js_Transmitt()  # intialisatiion des parametres par defaut   ### OKOKO
        if ("déploiement" in str(data['bouee_inst_act_bou']).lower()):
            js_transmi = fun_dcp_activ(data, ['bouee_modele', 'bouee_numero'],
                                       ['bouee_modele', 'bouee_numero'], '3', '3', dico_trams_oper,
                                       dico_trams, dico_trams_owner, allData, js_Transmitts)
            tab2_Transmitt.append(js_transmi)

        js_Transmitts = js_Transmitt()  # intialisatiion des parametres par defaut   ### OKOKO  Visite
        if ("renforcement" in str(data['bouee_inst_act_bou']).lower()) or ("visite" in str(data['bouee_inst_act_bou']).lower()):
            js_transmi = fun_dcp_activ(data, ['bouee_modele', 'bouee_numero'],
                                       ['bouee_modele', 'bouee_numero'], '1', '3', dico_trams_oper,
                                       dico_trams, dico_trams_owner, allData, js_Transmitts)
            tab2_Transmitt.append(js_transmi)

        js_Transmitts = js_Transmitt()  # intialisatiion des parametres par defaut  ### OKOKO
        if (("récupération" in str(data['bouee_inst_act_bou']).lower()) or ("perte" in str(data['bouee_inst_act_bou']).lower())):
            js_transmi = fun_dcp_activ(data, ['bouee_modele', 'bouee_numero'],
                                       ['bouee_modele', 'bouee_numero'], '2', '3', dico_trams_oper,
                                       dico_trams, dico_trams_owner, allData, js_Transmitts)
            tab2_Transmitt.append(js_transmi)


        js_Transmitts = js_Transmitt()  # intialisatiion des parametres par defaut    ### OKOKO
        if ("fin" in str(data['bouee_inst_act_bou']).lower()):
            js_transmi = fun_dcp_activ(data, ['bouee_modele', 'bouee_numero'],
                                       ['bouee_modele', 'bouee_numero'], '5', '3', dico_trams_oper,
                                       dico_trams, dico_trams_owner, allData, js_Transmitts)
            tab2_Transmitt.append(js_transmi)

    return tab2_Transmitt


def obj_ob_part_body_(temp_float, tab1_Float, js_Floats, bool_tuple=("false", "false")):
    """Fonction qui permet de traiter les materiels

    Args:
        tab1_Float (list): liste de données pour les materiels
        js_Floats (json): structure json pour un enregistrement sur les materiels

    Returns:

    """
    js_Floats = js_Float()
    if temp_float == None:
        pass
    elif type(temp_float) == tuple:

        js_Floats["objectMaterial"] = temp_float[0]
        js_Floats["whenArriving"] = bool_tuple[0]
        js_Floats["whenLeaving"] = bool_tuple[1]
        tab1_Float.append(js_Floats)

        js_Floats = js_Float()
        js_Floats["objectMaterial"] = temp_float[1]
        js_Floats["whenArriving"] = bool_tuple[0]
        js_Floats["whenLeaving"] = bool_tuple[1]
        tab1_Float.append(js_Floats)

    else:

        js_Floats["objectMaterial"] = temp_float
        js_Floats["whenArriving"] = bool_tuple[0]
        js_Floats["whenLeaving"] = bool_tuple[1]
        tab1_Float.append(js_Floats)


def errorFilter(response):
    """Permet de simplifier l'afficharge des erreurs dans le programme lors de l'insertion des données
    """
    error = json.loads(response)

    # print(error) ['exception']['result']['nodes']

    def errorFonction(nodes):
        if ('children' in nodes.keys()):
            return errorFonction(nodes['children'][0])

        if ('messages' in nodes.keys()):
            temp = nodes['messages']
            text = nodes['datum']['text']

            return "<strong>Texte : </strong>" + str(text) + "  <br>   <strong>Champs erreur: </strong>" + str(
                temp[0]['fieldName']) + " <br>  <strong>Message Erreur: </strong>" + str(temp[0]['message'])

    all_message = []

    if 'messages' in error['exception']['result']['nodes'][0].keys():
        all_message.append(errorFonction(error['exception']['result']['nodes'][0]))
    try:
        for val in error['exception']['result']['nodes'][0]['children']:
            all_message.append(errorFonction(val))
    except:
        pass

    return all_message


# # Supprimer un trip
# def del_trip(token, content):
#     dict = content
#     dicts = json.dumps(dict)

#     headers = {
#         "Content-Type": "application/json",
#         'authenticationToken': token
#     }

#     id_, ms_ = check_trip(token, content)

#     if ms_ == True:

#         id_ = id_.replace("#", "-")

#         url = 'https://observe.ob7.ird.fr/observeweb/api/public/data/ps/common/Trip/' + id_

#         print(id_)

#         print("Supprimer")

#         res = requests.delete(url, data=dicts, headers=headers)

#         print(res.status_code, "\n")

#         if res.status_code == 200:
#             print("Supprimer avec succes")
#             return json.loads(res.text)
#         else:
#             try:
#                 return errorFilter(res.text)
#             except KeyError:
#                 print("Message d'erreur: ", json.loads(res.text))


# Constitution de contents


# Si la premiere activité n'a pas de possition prendre la possition du port d'arriver

def build_trip(allData, info_bat, data_log, oce, prg, ob):
    """Fonction qui permet de contruire le gros fragment json de la marée et retourner des messages par rapport à la construction

    Args:
        allData (json): données de references
        info_bat (json): info sur le bateau date de depart/arrivée du port de depart/arrivé
        data_log (dataFrame):  les données du logbook
        oce (list): la liste des océans
        prg (list): pour la liste des programmes
        ob (json) : info sur le capitaine et homeid du bateau

    Returns:
        allMessages, js_contents
    """

    group = data_log.groupby(['date'])
    allMessages = []
    tab3_floatingObject = []
    activite = []
    routes = []
    ################## NEw
    js_catches = {}
    js_activitys = {}
    js_routeLogbooks = {}
    js_Transmitts = {}
    js_Floats = {}
    js_floatingObjects = {}
    #####################

    homeId = nb_r = Som_thon = 0
    nb = 1

    yft_id = getId(allData, "Species", argment="faoCode=YFT")
    skj_id = getId(allData, "Species", argment="faoCode=SKJ")
    bet_id = getId(allData, "Species", argment="faoCode=BET")
    germon_alb_id = getId(allData, "Species", argment="faoCode=ALB")
    WeightMeasureMet = getId(allData, "WeightMeasureMethod", argment="label2=Estimation visuelle")
    code_conser = getId(allData, "SpeciesFate", argment="code=6")
    code_conser_autre = getId(allData, "SpeciesFate", argment="code=15")
    code_reje = getId(allData, "SpeciesFate", argment="code=11")

    vers_code_6 = getId(allData, "VesselActivity", argment="code=6", domaine="seine")
    vers_code_13 = getId(allData, "VesselActivity", argment="code=13", domaine="seine")
    vers_code_21 = getId(allData, "VesselActivity", argment="code=21", domaine="seine")
    vers_code_99 = getId(allData, "VesselActivity", argment="code=99", domaine="seine")

    id_infoSource = getId(allData, "InformationSource", argment="code=S")
    id_dataQua = getId(allData, "DataQuality", argment="code=A")

    dico_code_sch_type = getAll(allData, "SchoolType")
    dico_code_setSucc = getAll(allData, "SetSuccessStatus")
    dico_objec = getAll(allData, "ObjectOperation")
    dico_trams_oper = getAll(allData, "TransmittingBuoyOperation")
    dico_trams = getAll(allData, "TransmittingBuoyType")
    dico_trams_owner = getAll(allData, "TransmittingBuoyOwnership")
    dico_objeMat = getAll(allData, "ObjectMaterial")
    tab_fpa = getAll(allData, "FpaZone", type_data="tableau")
    #############################

    oths = False
    oths_rej = []
    data_date = ""

    fpa_prece = ""
    heure_prece = ""
    comment_prece = ""
    nb_prece = 0
    not_time = False
    date = ""
    for val in group:
        #print(val)

        for index, data in val[1].iterrows():
            date = data["date"]

            # print(str(date).replace(" ","T").replace("00:00:00","")+str(data["heure"])+".000Z")
            tab4_catches = []

            # Permet d'incrementer le homeId sans perdre le file s'il n'y a pas d'activités
            if (data["lat1"] is not None) and (data["long1"] is not None) and (
                    data["zee"] is not None) and (
                    (data["cap_aut_esp_oth_esp"] is not None) or (data["cap_rej_dsc_esp"] is not None)):

                for vals in group:
                    for _, datas in vals[1].iterrows():
                        if data["date"] == datas["date"]:

                            if (datas['cap_alb_yft_p10_tail'] != None or datas['cap_alb_yft_p10_cap'] != None or \
                                    datas['cap_alb_yft_m10_tail'] != None or datas['cap_alb_yft_m10_cap'] != None or \
                                    datas['cap_lst_skj_tail'] != None or datas['cap_lst_skj_cap'] != None or \
                                    datas['cap_pat_bet_p10_tail'] != None or datas['cap_pat_bet_p10_cap'] != None or \
                                    datas['cap_pat_bet_m10_tail'] != None or datas['cap_pat_bet_m10_cap'] != None or \
                                    datas['cap_ger_alb_tail'] != None or datas['cap_ger_alb_cap'] != None or \
                                    datas['cap_aut_esp_oth_esp'] != None or datas['cap_aut_esp_oth_cap'] != None or \
                                    datas['cap_rej_dsc_esp'] != None or datas['cap_rej_dsc_cap'] != None):

                                ############################################  Modifier ##################################

                                def func_tab4_catches(js_catches, topId_sp, weight, WeightMeasureMet, code_conser):
                                    # js_catches["homeId"] = homeId
                                    js_catches["species"] = topId_sp
                                    js_catches["weight"] = weight
                                    js_catches["weightMeasureMethod"] = WeightMeasureMet
                                    js_catches["speciesFate"] = code_conser

                                    return js_catches

                                # Recuperation des caputure et chercher les infor a partir de l'api pour concevoir le content pour catches
                                js_catches = js_catche()  # intialisatiion des parametres defau
                                if (datas['cap_alb_yft_p10_tail'] != None and datas['cap_alb_yft_p10_cap'] != None):
                                    homeId += 1
                                    Som_thon += float(datas['cap_alb_yft_p10_cap'])

                                    js_catches = func_tab4_catches(js_catches, yft_id, datas['cap_alb_yft_p10_cap'],
                                                                   WeightMeasureMet, code_conser)
                                    tab4_catches.append(js_catches)

                                js_catches = js_catche()  # intialisatiion des parametres defau
                                if (datas['cap_alb_yft_m10_tail'] != None and datas['cap_alb_yft_m10_cap'] != None):
                                    homeId += 1
                                    Som_thon += float(datas['cap_alb_yft_m10_cap'])

                                    js_catches = func_tab4_catches(js_catches, yft_id, datas['cap_alb_yft_m10_cap'],
                                                                   WeightMeasureMet, code_conser)
                                    tab4_catches.append(js_catches)

                                js_catches = js_catche()  # intialisatiion des parametres defau
                                if (datas['cap_lst_skj_tail'] != None and datas['cap_lst_skj_cap'] != None):
                                    # print("Taille: ", datas['cap_lst_skj_tail'])

                                    homeId += 1
                                    Som_thon += float(datas['cap_lst_skj_cap'])

                                    js_catches = func_tab4_catches(js_catches, skj_id, datas['cap_lst_skj_cap'],
                                                                   WeightMeasureMet, code_conser)
                                    tab4_catches.append(js_catches)

                                js_catches = js_catche()  # intialisatiion des parametres defau
                                if (datas['cap_pat_bet_p10_tail'] != None and datas['cap_pat_bet_p10_cap'] != None):
                                    homeId += 1
                                    Som_thon += float(datas['cap_pat_bet_p10_cap'])

                                    js_catches = func_tab4_catches(js_catches, bet_id, datas['cap_pat_bet_p10_cap'],
                                                                   WeightMeasureMet, code_conser)
                                    tab4_catches.append(js_catches)

                                js_catches = js_catche()  # intialisatiion des parametres defau
                                if (datas['cap_pat_bet_m10_tail'] != None and datas['cap_pat_bet_m10_cap'] != None):
                                    homeId += 1
                                    Som_thon += float(datas['cap_pat_bet_m10_cap'])

                                    js_catches = func_tab4_catches(js_catches, bet_id, datas['cap_pat_bet_m10_cap'],
                                                                   WeightMeasureMet, code_conser)
                                    tab4_catches.append(js_catches)

                                js_catches = js_catche()  # intialisatiion des parametres defau
                                if (datas['cap_ger_alb_tail'] != None and datas['cap_ger_alb_cap'] != None):
                                    homeId += 1
                                    Som_thon += float(datas['cap_ger_alb_cap'])

                                    js_catches = func_tab4_catches(js_catches, germon_alb_id, datas['cap_ger_alb_cap'],
                                                                   WeightMeasureMet, code_conser)
                                    tab4_catches.append(js_catches)

                            ### Autre especes et rejets
                            # print("AAAA AAAAAA ", data["date"])
                            try:
                                if datas["cap_aut_esp_oth_esp"] is not None:
                                    if len(datas["cap_aut_esp_oth_esp"].replace(" ", "")) == 3:
                                        js_catches = js_catche()  # intialisatiion des parametres defau
                                        if (datas['cap_aut_esp_oth_esp'] != None and datas[
                                            'cap_aut_esp_oth_cap'] != None):
                                            homeId += 1
                                            if datas['cap_aut_esp_oth_tail'] == None:
                                                oths_id = getId(allData, "Species", argment="faoCode=" + (
                                                    datas["cap_aut_esp_oth_esp"]).upper())
                                                if (oths_id == "") or (oths_id == None):
                                                    oths_id = getId(allData, "Species", argment="faoCode=XXX*")
                                                    js_catches["comment"] = "Code espèce non trouvé: \"" + str(
                                                        datas["cap_aut_esp_oth_esp"]) + "\""

                                                # print("Esp trouv :  ", datas["cap_aut_esp_oth_esp"], oths_id)
                                                js_catches[
                                                    "species"] = oths_id  ###### PAscal  verifier si 3 lettres => recherce faoCode si trouve pas code recherche facode XXX* sinon XXX*
                                                js_catches["weight"] = datas['cap_aut_esp_oth_cap']
                                                Som_thon += float(datas[
                                                                      'cap_aut_esp_oth_cap'])  ####### Revenir sur la capture Pascal ####
                                                js_catches["weightMeasureMethod"] = WeightMeasureMet
                                                js_catches["speciesFate"] = code_conser_autre
                                            else:

                                                oths_id = getId(allData, "Species", argment="faoCode=" + (
                                                    datas["cap_aut_esp_oth_esp"]).upper())
                                                if (oths_id == "") or (oths_id == None):
                                                    oths_id = getId(allData, "Species", argment="faoCode=XXX*")
                                                    js_catches["comment"] = "Code espèce non trouvé: \"" + str(
                                                        datas["cap_aut_esp_oth_esp"]) + "\""
                                                # print("Esp trouv 2 :  ", datas["cap_aut_esp_oth_esp"], oths_id)
                                                Som_thon += float(datas['cap_aut_esp_oth_cap'])
                                                js_catches = func_tab4_catches(js_catches, oths_id,
                                                                               datas['cap_aut_esp_oth_cap'],
                                                                               WeightMeasureMet, code_conser_autre)

                                        tab4_catches.append(js_catches)
                                    else:
                                        allMessages.append("Le " + str(data["date"]) + " à " + str(data[
                                                                                                       "heure"]) + " ===> les espèces rejetées doivent être indiquées avec leur code FAO (ASFIS) 3 lettres. Le code trouvé est: " +
                                                           datas["cap_aut_esp_oth_esp"])

                                if datas["cap_rej_dsc_esp"] is not None:
                                    if len(datas["cap_rej_dsc_esp"].replace(" ", "")) == 3:
                                        js_catches = js_catche()  # intialisatiion des parametres defau
                                        if (datas['cap_rej_dsc_esp'] != None and datas['cap_rej_dsc_cap'] != None):
                                            homeId += 1
                                            if datas['cap_rej_dsc_tail'] == None:
                                                rejs_id = getId(allData, "Species",
                                                                argment="faoCode=" + (datas["cap_rej_dsc_esp"]).upper())
                                                if (rejs_id == "") or (rejs_id == None):
                                                    rejs_id = getId(allData, "Species", argment="faoCode=XXX*")
                                                    js_catches["comment"] = "Code espèce non trouvé: \"" + str(
                                                        datas["cap_rej_dsc_esp"]) + "\""
                                                js_catches[
                                                    "species"] = rejs_id  ###### PAscal  verifier si 3 lettres => recherce faoCode si trouve pas code recherche facode XXX* sinon XXX*
                                                js_catches["weight"] = datas['cap_rej_dsc_cap']
                                                Som_thon += float(datas[
                                                                      'cap_rej_dsc_cap'])  ####### Revenir sur la capture Pascal ####
                                                js_catches["weightMeasureMethod"] = WeightMeasureMet
                                                js_catches["speciesFate"] = code_reje
                                            else:
                                                rejs_id = getId(allData, "Species",
                                                                argment="faoCode=" + (datas["cap_rej_dsc_esp"]).upper())
                                                if (rejs_id == "") or (rejs_id == None):
                                                    rejs_id = getId(allData, "Species", argment="faoCode=XXX*")
                                                    js_catches["comment"] = "Code espèce non trouvé: \"" + str(
                                                        datas["cap_rej_dsc_esp"]) + "\""
                                                Som_thon += float(datas['cap_rej_dsc_cap'])
                                                js_catches = func_tab4_catches(js_catches, rejs_id,
                                                                               datas['cap_rej_dsc_cap'],
                                                                               WeightMeasureMet, code_reje)

                                        tab4_catches.append(js_catches)
                                    else:
                                        allMessages.append("Le " + str(data["date"]) + " à " + str(data[
                                                                                                       "heure"]) + " ===> les espèces rejetées doivent être indiquées avec leur code FAO (ASFIS) 3 lettres. Le code trouvé est: " +
                                                           datas["cap_rej_dsc_esp"])
                            except:
                                # Date heure ==> Les données concernant les espèces rejetées sont mal formatées (raison exacte indeterminée)
                                allMessages.append("Le " + str(data["date"]) + " à " + str(data[
                                                                                               "heure"]) + " ===> Les données concernant les espèces rejetées sont mal formatées (raison exacte indeterminée)")

            ################# Floating Obj ##############

            tab1_Float = []
            tab2_Transmitt = []
            temp_float = None

            def func_tab3_floatingObject(allData, data, dico_objeMat, js_Floats, bool_tuple, argment, tab1_Float=[]):

                # Types objets flottants
                js_Floats = js_Float()  # intialisatiion des parametres defau
                temp_float = floatingObjectPart(data['obj_flot_typ_obj'], data, dico_objeMat, index='obj_flot_typ_obj')

                obj_ob_part_body_(temp_float, tab1_Float, js_Floats, bool_tuple)

                # Risque de maillage en surface
                js_Floats = js_Float()  # intialisatiion des parametres defau
                temp_float = floatingObjectPart(data['obj_flot_risq_mail_en_surf'], data, dico_objeMat,
                                                index='obj_flot_risq_mail_en_surf')
                obj_ob_part_body_(temp_float, tab1_Float, js_Floats, bool_tuple)

                # Risque de maillage sous la surface
                js_Floats = js_Float()  # intialisatiion des parametres defau
                temp_float = floatingObjectPart(data['obj_flot_risq_mail_sou_surf'], data, dico_objeMat,
                                                index='obj_flot_risq_mail_sou_surf', perte_act=True)
                obj_ob_part_body_(temp_float, tab1_Float, js_Floats, bool_tuple)

                js_floatingObjects = js_floatingObject(tab2_Transmitt, tab1_Float)
                js_floatingObjects["objectOperation"] = getId(allData, "ObjectOperation", argment)

                return js_floatingObjects

            try:
                if (data['obj_flot_act_sur_obj'] == None) or ("perte" in str(data['bouee_inst_act_bou']).lower()):
                    depart_date = info_bat['Depart_Date']

                    if (str(data["date"]).split(" ")[0] >= depart_date) and (
                            "perte" in str(data['bouee_inst_act_bou']).lower()):
                        # print("perte")
                        operation = "perte"
                        tab2_Transmitt = obj_deja_deploy(data, js_Transmitts, dico_trams_oper, dico_trams,
                                                         dico_trams_owner, allData, operation)
                        # print("perte ou Fin ", data['date'])
                        # code 13 mettre avec ce que
                        js_floatingObjects = func_tab3_floatingObject(allData, data, dico_objeMat, js_Float,
                                                                      bool_tuple=("true", "true"), argment="code=11")
                        tab3_floatingObject.append(js_floatingObjects)

                elif ("mise" in str(data['obj_flot_act_sur_obj']).lower()):
                    operation = "mise à l'eau"
                    tab2_Transmitt = obj_deja_deploy(data, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner,
                                                     allData, operation)
                    # print("Mise ", data['date'])
                    js_floatingObjects = func_tab3_floatingObject(allData, data, dico_objeMat, js_Float,
                                                                  bool_tuple=("false", "true"), argment="code=1")
                    tab3_floatingObject.append(js_floatingObjects)

                elif ("visite" in str(data['obj_flot_act_sur_obj']).lower()):
                    operation = "visite"
                    # print(data)
                    tab2_Transmitt = obj_deja_deploy(data, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner,
                                                     allData, operation)
                    # print("visite ", data['date'])
                    js_floatingObjects = func_tab3_floatingObject(allData, data, dico_objeMat, js_Float,
                                                                  bool_tuple=("true", "false"), argment="code=2")
                    tab3_floatingObject.append(js_floatingObjects)

                elif ("renforcement" in str(data['obj_flot_act_sur_obj']).lower()):
                    operation = "renforcement"
                    tab2_Transmitt = obj_deja_deploy(data, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner,
                                                     allData, operation)
                    # print("renforcement ", data['date'])
                    js_floatingObjects = func_tab3_floatingObject(allData, data, dico_objeMat, js_Float,
                                                                  bool_tuple=("true", "true"), argment="code=8")
                    tab3_floatingObject.append(js_floatingObjects)

                elif ("retrait" in str(data['obj_flot_act_sur_obj']).lower()):
                    operation = "retrait"
                    tab2_Transmitt = obj_deja_deploy(data, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner,
                                                     allData, operation)
                    # print("retrat ", data['date'])
                    js_floatingObjects = func_tab3_floatingObject(allData, data, dico_objeMat, js_Float,
                                                                  bool_tuple=("true", "false"), argment="code=4")
                    tab3_floatingObject.append(js_floatingObjects)

                elif (("perte" in str(data['obj_flot_act_sur_obj']).lower()) or (
                        "fin" in str(data['obj_flot_act_sur_obj']).lower())):
                    operation = "Perte ou de fin d'utilisation"
                    tab2_Transmitt = obj_deja_deploy(data, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner,
                                                     allData, operation)
                    # print("perte ou Fin ", data['date'])
                    js_floatingObjects = func_tab3_floatingObject(allData, data, dico_objeMat, js_Float,
                                                                  bool_tuple=("true", "true"), argment="code=11")
                    tab3_floatingObject.append(js_floatingObjects)

                if ((data['obj_flot_act_sur_obj'] == None) and (data['obj_flot_typ_obj'] != None) and (
                        "perte" != str(data['bouee_inst_act_bou']).lower())):
                    allMessages.append("Le " + str(data["date"]) + " à " + str(
                        data["heure"]) + " ===> Activité sur objet flottant non renseignéé ")

            except TransmitException as e:
                # print(data["date"],data["heure"]," Logbook non conforme #######")
                allMessages.append(e.message)
            ##############################################################################################################################################

            ########### Activite ############

            depart_date = info_bat['Depart_Date']
            Depart_heure = info_bat['Depart_heure']

            last = len(data_log) - 1
            if (((data["heure"] is not None) or (data["heure"] is None)) and (data["lat1"] is not None) and (
                    data["long1"] is not None) and (
                        data["zee"] is not None)) or (((str(data["date"]).split(" ")[0] == depart_date) and (
                    str(Depart_heure) == str(data["heure"]))) or (index == last)):

                if data["heure"] is None:
                    not_time = True
                js_activitys = js_activity(tab4_catches, tab3_floatingObject)

                # print("BBBBBB BBB ", data["date"],"   ", data["heure"],"   ", data["calee_porta"],"   ", data["calee_nul"])

                ######################### NEW LIGnE
                tab3_floatingObject = []
                ###########################

                if tab4_catches is not []:
                    if data["comment"] is not None:
                        js_activitys["comment"] = data["comment"]
                    else:
                        js_activitys["comment"] = "Aucun commentaire"
                else:
                    js_activitys["comment"] = "Le programme de lecture du livre de bord n’a pas pu \
                                                déterminer le code bateau. Le nom du bateau était " + info_bat['Navire']
                if data["heure"] is None:
                    js_activitys["time"] = data["heure"]
                else:
                    js_activitys["time"] = str(date).replace(" ", "T").replace("00:00:00", "") + str(
                        data["heure"]) + ".000Z"
                js_activitys["seaSurfaceTemperature"] = data["temp_mer"]
                js_activitys["windDirection"] = data["vent_dir"]

                if Som_thon != 0:
                    js_activitys["totalWeight"] = Som_thon

                # Verifier si premiere activité et enregistrer
                if ((str(data["date"]).split(" ")[0] == depart_date) and (str(Depart_heure) == str(data["heure"]))):
                    if data["lat1"] != None and data["lat2"] != None and \
                            data["lat3"] != None and data["long1"] != None and \
                            data["long2"] != None and data["long3"] != None:
                        js_activitys["latitude"], js_activitys["longitude"], checkMsg = lat_long(data["lat1"],
                                                                                                 data["lat2"],
                                                                                                 data["lat3"],
                                                                                                 data["long1"],
                                                                                                 data["long2"],
                                                                                                 data["long3"])

                        returnMsg = "Position manquante (latitude / longitude): Le " + str(data["date"]).replace(
                            " 00:00:00", " à ") + str(data["heure"])
                        if checkMsg:
                            allMessages.append(returnMsg)
                    else:
                        js_activitys["latitude"], js_activitys["longitude"] = get_lat_long(allData,
                                                                                           info_bat['Depart_Port'])

                elif index == last:
                    if data["lat1"] != None and data["lat2"] != None and \
                            data["lat3"] != None and data["long1"] != None and \
                            data["long2"] != None and data["long3"] != None:
                        js_activitys["latitude"], js_activitys["longitude"], checkMsg = lat_long(data["lat1"],
                                                                                                 data["lat2"],
                                                                                                 data["lat3"],
                                                                                                 data["long1"],
                                                                                                 data["long2"],
                                                                                                 data["long3"])

                        returnMsg = "Position manquant (latitude / longitude): Le " + str(data["date"]).replace(
                            " 00:00:00", " à ") + str(data["heure"])
                        if checkMsg:
                            allMessages.append(returnMsg)
                    else:
                        js_activitys["latitude"], js_activitys["longitude"] = get_lat_long(allData,
                                                                                           info_bat['Arrivee_Port'])
                else:
                    js_activitys["latitude"], js_activitys["longitude"], checkMsg = lat_long(data["lat1"], data["lat2"],
                                                                                             data["lat3"],
                                                                                             data["long1"],
                                                                                             data["long2"],
                                                                                             data["long3"])

                    returnMsg = "Position manquant (latitude / longitude): Le " + str(data["date"]).replace(" 00:00:00",
                                                                                                            " à ") + str(
                        data["heure"])
                    if checkMsg:
                        allMessages.append(returnMsg)

                def schoolType(chaine, dico_code_sch_type):
                    if chaine != None and chaine != "":
                        if ("libre" in chaine.lower()):
                            ## Table SchoolType code 2
                            return dico_code_sch_type["2"]

                        elif ("objet" in chaine.lower()):
                            ## Table SchoolType code 1
                            return dico_code_sch_type["1"]

                        elif ("ind" in chaine.lower()):
                            ## Table SchoolType code 0
                            return dico_code_sch_type["0"]

                def setCo_setSuc_vess(setCount, setSuccessStatus, vesselActivity):
                    return setCount, setSuccessStatus, vesselActivity

                messg = "Le " + str(data["date"]).replace(" 00:00:00", " à ") + str(
                    data["heure"]) + " ===> Le type de banc n'a pas été précisé"

                if (data['calee_porta'] is not None) and (data['calee_nul'] is None):
                    # Code 6
                    js_activitys["setCount"], js_activitys["setSuccessStatus"], js_activitys[
                        "vesselActivity"] = setCo_setSuc_vess(1, dico_code_setSucc["1"], vers_code_6)

                    if data["calee_type"] is None:
                        allMessages.append(messg)
                    else:
                        js_activitys["schoolType"] = schoolType(data["calee_type"], dico_code_sch_type)


                elif (data['calee_porta'] is None) and (data['calee_nul'] is not None):
                    # Code 6
                    js_activitys["setCount"], js_activitys["setSuccessStatus"], js_activitys[
                        "vesselActivity"] = setCo_setSuc_vess(1, dico_code_setSucc["0"], vers_code_6)

                    if data["calee_type"] is None:
                        allMessages.append(messg)
                    else:
                        js_activitys["schoolType"] = schoolType(data["calee_type"], dico_code_sch_type)


                elif (data['calee_porta'] is not None) and (data['calee_nul'] is not None):
                    # Code 6
                    js_activitys["setCount"], js_activitys["setSuccessStatus"], js_activitys[
                        "vesselActivity"] = setCo_setSuc_vess(1, dico_code_setSucc["2"], vers_code_6)

                    if data["calee_type"] is None:
                        allMessages.append(messg)
                    else:
                        js_activitys["schoolType"] = schoolType(data["calee_type"], dico_code_sch_type)

                elif data['obj_flot_act_sur_obj'] is not None:
                    # Code 13
                    js_activitys["setCount"], js_activitys["setSuccessStatus"], js_activitys[
                        "vesselActivity"] = setCo_setSuc_vess(0, None, vers_code_13)
                    js_activitys["schoolType"] = None

                else:
                    if (str(data["date"]).split(" ")[0] >= depart_date) and (
                            "perte" in str(data['bouee_inst_act_bou']).lower()):
                        # Code 13
                        js_activitys["setCount"], js_activitys["setSuccessStatus"], js_activitys[
                            "vesselActivity"] = setCo_setSuc_vess(0, None, vers_code_13)
                        js_activitys["schoolType"] = None
                    elif ("perte" not in str(data['bouee_inst_act_bou']).lower()):
                        # Code 99
                        js_activitys["setCount"], js_activitys["setSuccessStatus"], js_activitys[
                            "vesselActivity"] = setCo_setSuc_vess(0, None, vers_code_99)
                        js_activitys["schoolType"] = None

                js_activitys["informationSource"] = id_infoSource
                js_activitys["dataQuality"] = id_dataQua
                js_activitys["currentFpaZone"], comment_temp = fpaZone_id(data["zee"], tab_fpa, allData)
                if (comment_temp != "") and (comment_temp != None):
                    js_activitys["comment"] = js_activitys["comment"] + " # " + comment_temp

                ObserSys0 = getSome(allData, argment="code=0", moduleName="ObservedSystem")[0]["topiaId"]
                observedSystem = [ObserSys0]

                if data["asso_bc_libre"] != None:
                    observedSystem.remove(ObserSys0)
                    observedSystem.append(getSome(allData, argment="code=0", moduleName="ObservedSystem")[0]["topiaId"])

                elif data["asso_objet"] != None:
                    observedSystem.remove(ObserSys0)
                    observedSystem.append(
                        getSome(allData, argment="code=20", moduleName="ObservedSystem")[0]["topiaId"])

                elif data["asso_balise"] != None:
                    observedSystem.remove(ObserSys0)
                    observedSystem.append(
                        getSome(allData, argment="code=20", moduleName="ObservedSystem")[0]["topiaId"])

                elif data["asso_baliseur"] != None:
                    observedSystem.remove(ObserSys0)
                    observedSystem.append(
                        getSome(allData, argment="code=28", moduleName="ObservedSystem")[0]["topiaId"])

                elif data["asso_requin"] != None:
                    observedSystem.remove(ObserSys0)
                    observedSystem.append(
                        getSome(allData, argment="code=12", moduleName="ObservedSystem")[0]["topiaId"])

                elif data["asso_baleine"] != None:
                    observedSystem.remove(ObserSys0)
                    observedSystem.append(
                        getSome(allData, argment="code=11", moduleName="ObservedSystem")[0]["topiaId"])

                elif data["asso_oiseaux"] != None:
                    observedSystem.remove(ObserSys0)
                    observedSystem.append(getSome(allData, argment="code=4", moduleName="ObservedSystem")[0]["topiaId"])

                js_activitys["observedSystem"] = observedSystem
                js_activitys["wind"] = get_wind_id_interval(allData, "Wind", data["vent_vit"])

                js_activitys["number"] = int(nb)

                if (heure_prece == data["heure"]) and (date_prece == data["date"]):
                    # Supprimer activité inseré en dernier
                    if len(activite) != 0:
                        del (activite[-1])
                    # faire le changement de zone
                    js_activitys["number"] = nb_prece
                    js_activitys["previousFpaZone"], comment_temp = fpaZone_id(fpa_prece, tab_fpa, allData)
                    js_activitys["nextFpaZone"], comment_temp = fpaZone_id(data["zee"], tab_fpa, allData)
                    js_activitys["currentFpaZone"] = None

                    if (comment_prece is not None) and (
                            (js_activitys["comment"] is not None) and (js_activitys["comment"] != "Aucun commentaire")):
                        js_activitys["comment"] = comment_prece + " ==> " + js_activitys["comment"]
                    elif (comment_prece is None) and (js_activitys["comment"] is not None):
                        js_activitys["comment"] = js_activitys["comment"]
                    elif (comment_prece is not None) and (
                            (js_activitys["comment"] is None) or (js_activitys["comment"] == "Aucun commentaire")):
                        js_activitys["comment"] = comment_prece
                    else:
                        js_activitys["comment"] = "Aucun commentaire"

                    if (comment_temp != "") and (comment_temp != None):
                        js_activitys["comment"] = js_activitys["comment"] + " # " + comment_temp

                    js_activitys["vesselActivity"] = vers_code_21

                    # ajouter la nouvelle activité
                    activite.append(js_activitys)
                    # Reinitialliser les variables
                    nb = nb_prece
                else:
                    activite.append(js_activitys)

                fpa_prece = data["zee"]
                heure_prece = data["heure"]
                date_prece = data["date"]
                comment_prece = data["comment"]
                nb_prece = int(nb)
                Som_thon = 0
                nb += 1

        js_routeLogbooks = js_routeLogbook(activite)

        # print("DDDDDD DDDDD DDDD", type(data["heure"]))

        js_routeLogbooks["date"] = str(date).replace(" ", "T") + ".000Z"

        routes.append(js_routeLogbooks)

        activite = []

        nb = 1
        nb_r += 1

    #print("Yes", routes)

    js_contents = js_content(routes, oce, prg)
    # activitiesAcquisitionMode [BY_NUMBER, BY_TIME]
    if not_time:
        js_contents["activitiesAcquisitionMode"] = "BY_NUMBER"
    else:
        js_contents["activitiesAcquisitionMode"] = "BY_TIME"
    # noinspection PyBroadException
    try:
        # si plusieurs Rechercher celui qui a le code le plus elévé avec toujours son status == 1
        #########################################################""
        js_contents["vessel"] = getId(allData, "Vessel",
                                      argment="label2=" + info_bat['Navire'] + "&filters.status=enabled", nbArg=3)
        if js_contents["vessel"] is None:
            js_contents["vessel"] = getId(allData, "Vessel",
                                          argment="label2=" + info_bat['Navire'].upper() + "&filters.status=enabled",
                                          nbArg=3)
        if js_contents["vessel"] is None:
            js_contents["vessel"] = getId(allData, "Vessel",
                                          argment="label2=" + info_bat['Navire'].lower() + "&filters.status=enabled",
                                          nbArg=3)
        if js_contents["vessel"] is None:
            allMessages.append(
                "Le nom du Navire n'a pas été trouvé. Veuillez verifier son existance dans la base de données puis corriger dans le livre de bord.")
    except:
        pass

    if info_bat['Depart_Port'] == None:
        js_contents["departureHarbour"] = getId(allData, "Harbour", argment="code=999")
    else:
        resu = getId(allData, "Harbour", argment="label2=" + (info_bat['Depart_Port']).upper())
        if resu == None:
            js_contents["departureHarbour"] = getId(allData, "Harbour", argment="code=999")
        else:
            js_contents["departureHarbour"] = resu

    if info_bat['Arrivee_Port'] == None:
        js_contents["landingHarbour"] = getId(allData, "Harbour", argment="code=999")
    else:
        resu = getId(allData, "Harbour", argment="label2=" + (info_bat['Arrivee_Port']).upper())
        if resu == None:
            js_contents["landingHarbour"] = getId(allData, "Harbour", argment="code=999")
        else:
            js_contents["landingHarbour"] = resu

    if info_bat['Depart_Date'] == None:
        js_contents["startDate"] = None
    else:
        js_contents["startDate"] = info_bat['Depart_Date'] + "T00:00:00.000Z"  # "2021-03-02T00:00:00.000Z" #

    if info_bat['Arrivee_Date'] == None:
        js_contents["endDate"] = None
    else:
        js_contents["endDate"] = info_bat['Arrivee_Date'] + "T00:00:00.000Z"

    js_contents["captain"], js_contents["logbookDataEntryOperator"] = cap_obs_sea(allData, ob)

    js_contents["loch"] = info_bat['Arrivee_Loch']
    js_contents["homeId"] = str(ob["mar_homeId"])
    js_contents["observationsProgram"] = None
    js_contents['logbookComment'] = "NB: Service Web"

    return allMessages, js_contents


def build_trip_v23(allData, info_bat, data_log, oce, prg):
    """Fonction qui permet de contruire le gros fragment json de la marée et retourner des messages par rapport à la construction

    Args:
        allData (json): données de references
        info_bat (json): info sur le bateau date de depart/arrivée du port de depart/arrivé et info sur le capitaine et homeid du bateau
        data_log (dataFrame):  les données du logbook
        oce (list): la liste des océans
        prg (list): pour la liste des programmes

    Returns:
        allMessages, js_contents
    """

    group = data_log.groupby(['date'])
    oths = False
    oths_rej = []
    data_date = ""

    fpa_suiv = ""
    type_decla_suiv = ""
    comment_suiv = ""
    comment_prece = ""
    nb_prece = 0
    not_time = False

    allMessages = []
    tab3_floatingObject = []
    activite = []
    routes = []
    ################## NEw
    js_catches = {}
    js_activitys = {}
    js_routeLogbooks = {}
    js_Transmitts = {}
    js_Floats = {}
    js_floatingObjects = {}
    #####################

    homeId = Som_thon = 0
    nb = 1

    WeightMeasureMet = getId(allData, "WeightMeasureMethod", argment="label2=Estimation visuelle")

    code_conser = getId(allData, "SpeciesFate", argment="code=6")
    code_conser_autre = getId(allData, "SpeciesFate", argment="code=15")
    code_reje = getId(allData, "SpeciesFate", argment="code=11")

    # Champs type de declaration dans le logbook
    vers_code_0 = getId(allData, "VesselActivity", argment="code=0", domaine="seine")
    vers_code_6 = getId(allData, "VesselActivity", argment="code=6", domaine="seine")
    vers_code_13 = getId(allData, "VesselActivity", argment="code=13", domaine="seine")
    vers_code_21 = getId(allData, "VesselActivity", argment="code=21", domaine="seine")
    vers_code_99 = getId(allData, "VesselActivity", argment="code=99", domaine="seine")
    vers_code_103 = getId(allData, "VesselActivity", argment="code=103", domaine="seine")

    id_infoSource = getId(allData, "InformationSource", argment="code=S")
    id_dataQua = getId(allData, "DataQuality", argment="code=A")

    dico_code_sch_type = getAll(allData, "SchoolType")
    dico_code_setSucc = getAll(allData, "SetSuccessStatus")
    dico_objec = getAll(allData, "ObjectOperation")
    dico_trams_oper = getAll(allData, "TransmittingBuoyOperation")
    dico_trams = getAll(allData, "TransmittingBuoyType")
    dico_trams_owner = getAll(allData, "TransmittingBuoyOwnership")
    dico_objeMat = getAll(allData, "ObjectMaterial")

    tab_fpa = getAll(allData, "FpaZone", type_data="tableau")
    #############################



    i = 0
    for val in group:
        i += 1

        Som_thon = 0
        list_catches = []
        # nb += 1

        number = 1
        inner = val[1].groupby(['date', 'heure'])

        for vals in inner:
            for index, datas in vals[1].iterrows():
                data = datas
                date = data["date"]

                if (data['espece'] != None and (data['quant_conser_tonne'] != None or data['quant_reje_tonne'] != None)):

                    def func_tab4_catches(js_catches, topId_sp, weight, WeightMeasureMet, code_conser_reje, count=None, wgtCategory=None):
                        js_catches["species"] = topId_sp
                        js_catches["weight"] = weight
                        js_catches["weightMeasureMethod"] = WeightMeasureMet
                        js_catches["speciesFate"] = code_conser_reje
                        js_catches["count"] = count
                        js_catches["weightCategory"] = wgtCategory

                        return js_catches

                    wgtCategory = weightCategory(allData, data['categ_poids'], data["espece"][:3].upper())

                    species_id = getId(allData, "Species", argment="faoCode=" + data["espece"][:3].upper())
                    if species_id == "":
                        species_id = getId(allData, "Species", argment="faoCode=XXX")

                    # Recuperation des caputure et chercher les infor a partir de l'api pour concevoir le content pour catches
                    js_catches = js_catche()  # intialisatiion des parametres defaut

                    homeId += 1

                    if data['quant_conser_tonne'] != None:

                        Som_thon += float(data['quant_conser_tonne']) # Tonne
                        js_catches = func_tab4_catches(js_catches, species_id, data['quant_conser_tonne'],
                                                       WeightMeasureMet, code_conser, data['quant_conser_nb'], wgtCategory)
                    elif data['quant_reje_tonne'] != None:

                        Som_thon += float(data['quant_reje_tonne']) # Tonne
                        js_catches = func_tab4_catches(js_catches, species_id, data['quant_reje_tonne'],
                                                       WeightMeasureMet, code_reje, data['quant_reje_nb'], wgtCategory)
                    else:
                        Som_thon += float(0) # Tonne
                        js_catches = func_tab4_catches(js_catches, species_id, None,
                                                       WeightMeasureMet, code_conser, data['quant_conser_nb'], wgtCategory)

                    list_catches.append(js_catches)

            ##### Floating Obj CODE ###########
            d_date  = vals[1].iloc[0]['date']
            d_act_boue  = vals[1].iloc[0]['bouee_inst_act_bou']

            depart_date = info_bat['Depart_Date']
            Depart_heure = info_bat['Depart_heure']

            # recuperer les objets
            data_act_obj = vals[1].loc[:, 'obj_flot_act_sur_obj':'obj_bio']
            data_act_bo = vals[1].loc[:, 'bouee_inst_act_bou':'bouee_numero']

            # supprimer les doublons
            d_act_obj = data_act_obj.drop_duplicates()
            d_act_bo = data_act_bo.drop_duplicates()


            check_vis_dep = ()
            for index, datas in d_act_obj.iterrows():

                data = datas
                tab1_Float = []
                tab2_Transmitt = []
                temp_float = None

                def func_tab3_floatingObject(allData, data, dico_objeMat, js_Floats, bool_tuple, argment, tab1_Float=[]):

                    # Types objets flottants
                    js_Floats = js_Float()  # intialisatiion des parametres defaut
                    temp_float = floatingObjectPart(data['obj_flot_typ_obj'], data, dico_objeMat, index='obj_flot_typ_obj')
                    obj_ob_part_body_(temp_float, tab1_Float, js_Floats, bool_tuple)

                    try:
                        # Maillage
                        js_Floats = js_Float()  # intialisatiion des parametres defau
                        temp_float = floatingObjectPart(data['obj_mailles'], data, dico_objeMat,
                                                        index='obj_mailles')
                        obj_ob_part_body_(temp_float, tab1_Float, js_Floats, bool_tuple)

                    except:
                        # Risque de maillage en surface
                        js_Floats = js_Float()  # intialisatiion des parametres defau
                        temp_float = floatingObjectPart(data['obj_flot_risq_mail_en_surf'], data, dico_objeMat,
                                                        index='obj_flot_risq_mail_en_surf')
                        obj_ob_part_body_(temp_float, tab1_Float, js_Floats, bool_tuple)

                        # Risque de maillage sous la surface
                        js_Floats = js_Float()  # intialisatiion des parametres defau
                        temp_float = floatingObjectPart(data['obj_flot_risq_mail_sou_surf'], data, dico_objeMat,
                                                        index='obj_flot_risq_mail_sou_surf', perte_act=True)
                        obj_ob_part_body_(temp_float, tab1_Float, js_Floats, bool_tuple)

                    js_floatingObjects = js_floatingObject(tab2_Transmitt, tab1_Float)
                    js_floatingObjects["objectOperation"] = getId(allData, "ObjectOperation", argment)

                    return js_floatingObjects


                prev = -1

                try:
                    check_vis_dep = tuple()
                    if (data['obj_flot_act_sur_obj'] != None) and (d_act_obj.loc[index + 1, 'obj_flot_act_sur_obj'] != None):
                        check_vis_dep = data['obj_flot_act_sur_obj'].lower(), d_act_obj.loc[index + 1, 'obj_flot_act_sur_obj'].lower()
                        prev = index + 1

                        if ("visite" in check_vis_dep) and ("déploiement" in check_vis_dep):

                            operation = "renforcement == Visite + Déploiement"
                            tab2_Transmitt = obj_deja_deploy_v23(d_act_bo, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner,
                                                             allData, operation)
                            js_floatingObjects = func_tab3_floatingObject(allData, d_act_obj, dico_objeMat, js_Float,
                                                                          bool_tuple=("true", "true"), argment="code=8")
                            tab3_floatingObject.append(js_floatingObjects)
                except:
                    if index != prev:
                        try:

                            if (data['obj_flot_act_sur_obj'] == None) or ("perte" in str(d_act_boue).lower()):
                                depart_date = info_bat['Depart_Date']

                                if (str(d_date).split(" ")[0] >= depart_date) and (
                                        "perte" in str(d_act_boue).lower()):   ##### A corriger
                                    operation = "perte"
                                    tab2_Transmitt = obj_deja_deploy_v23(d_act_bo, js_Transmitts, dico_trams_oper, dico_trams,
                                                                     dico_trams_owner, allData, operation)  ############# A corriger
                                    js_floatingObjects = func_tab3_floatingObject(allData, d_act_obj, dico_objeMat, js_Float,
                                                                                  bool_tuple=("true", "true"), argment="code=11")
                                    tab3_floatingObject.append(js_floatingObjects)

                            elif ("déploiement" in str(data['obj_flot_act_sur_obj']).lower()):  ####### Déploiement
                                operation = "mise à l'eau == Déploiement"
                                tab2_Transmitt = obj_deja_deploy_v23(d_act_bo, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner,
                                                                 allData, operation)
                                js_floatingObjects = func_tab3_floatingObject(allData, d_act_obj, dico_objeMat, js_Float,
                                                                              bool_tuple=("false", "true"), argment="code=1")
                                tab3_floatingObject.append(js_floatingObjects)

                            elif ("visite" in str(data['obj_flot_act_sur_obj']).lower()):  ###### Visite
                                operation = "visite == Visite"
                                tab2_Transmitt = obj_deja_deploy_v23(d_act_bo, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner,
                                                                 allData, operation)
                                js_floatingObjects = func_tab3_floatingObject(allData, d_act_obj, dico_objeMat, js_Float,
                                                                              bool_tuple=("true", "false"), argment="code=2")
                                tab3_floatingObject.append(js_floatingObjects)

                            elif ("pêche" in str(data['obj_flot_act_sur_obj']).lower()):  ###### Pêche
                                operation = "pêche == Pêche"
                                tab2_Transmitt = obj_deja_deploy_v23(d_act_bo, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner,
                                                                 allData, operation)
                                js_floatingObjects = func_tab3_floatingObject(allData, d_act_obj, dico_objeMat, js_Float,
                                                                              bool_tuple=("true", "false"), argment="code=6")
                                tab3_floatingObject.append(js_floatingObjects)


                            elif ("récupération" in str(data['obj_flot_act_sur_obj']).lower()): ####### Récupération
                                operation = "retrait == Récupération"
                                tab2_Transmitt = obj_deja_deploy_v23(d_act_bo, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner,
                                                                 allData, operation)
                                js_floatingObjects = func_tab3_floatingObject(allData, d_act_obj, dico_objeMat, js_Float,
                                                                              bool_tuple=("true", "false"), argment="code=4")
                                tab3_floatingObject.append(js_floatingObjects)

                            elif (("perte" in str(data['obj_flot_act_sur_obj']).lower()) or (
                                    "fin" in str(data['obj_flot_act_sur_obj']).lower())):  ####### Perte ou Fin d'utilisation
                                operation = "Perte ou de fin d'utilisation == Perte ou Fin d'utilisation"
                                tab2_Transmitt = obj_deja_deploy_v23(d_act_bo, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner,
                                                                 allData, operation)
                                js_floatingObjects = func_tab3_floatingObject(allData, d_act_obj, dico_objeMat, js_Float,
                                                                              bool_tuple=("true", "true"), argment="code=11")
                                tab3_floatingObject.append(js_floatingObjects)


                            if ((data['obj_flot_act_sur_obj'] == None) and (data['obj_flot_typ_obj'] != None) and ("perte" != str(data['bouee_inst_act_bou']).lower())):
                                allMessages.append("Le " + str(data["date"]) + " à " + str(data["heure"]) + " ===> Activité sur objet flottant non renseignéé ")

                        except TransmitException as e:
                            allMessages.append(e.message)



                    elif ("déploiement" in str(data['obj_flot_act_sur_obj']).lower()):
                        operation = "renforcement == Visite + Déploiement"
                        tab2_Transmitt = obj_deja_deploy_v23(d_act_bo, js_Transmitts, dico_trams_oper, dico_trams, dico_trams_owner,
                                                         allData, operation)
                        js_floatingObjects = func_tab3_floatingObject(allData, d_act_obj, dico_objeMat, js_Float,
                                                                              bool_tuple=("true", "true"), argment="code=8")
                        tab3_floatingObject.append(js_floatingObjects)

            ########### Activite ############
            data_activity  = vals[1].iloc[0]

            last = len(data_log) - 1

            if data_activity["heure"] is None:
                not_time = True

            js_activitys = js_activity(list_catches, tab3_floatingObject)

            tab3_floatingObject = []

            if list_catches is not []:

                if data_activity["commentaire"] is not None:
                    js_activitys["comment"] = data_activity["commentaire"]
                else:
                    js_activitys["comment"] = "Aucun commentaire"
            else:
                js_activitys["comment"] = "Le programme de lecture du livre de bord n’a pas pu \
                                            déterminer le code bateau. Le nom du bateau était " + info_bat['Navire']
            if data_activity["heure"] is None:
                js_activitys["time"] = data_activity["heure"]
            else:
                js_activitys["time"] = str(date).replace(" ", "T").replace("00:00:00", "") + str(
                data_activity["heure"]) + ".000Z"
            js_activitys["seaSurfaceTemperature"] = data_activity["temp_mer"]
            js_activitys["windDirection"] = data_activity["vent_dir"]

            if Som_thon != 0:
                js_activitys["totalWeight"] = Som_thon


            # Verifier si premiere activité et enregistrer
            if ( ((str(data_activity["date"]).split(" ")[0] == depart_date) and (str(Depart_heure) == str(data_activity["heure"]))) or ("DEP" in data_activity["type_declaration"].upper()) ):
                if data_activity["lat1"] != None and data_activity["lat2"] != None and \
                        data_activity["lat3"] != None and data_activity["long1"] != None and \
                        data_activity["long2"] != None and data_activity["long3"] != None:
                    js_activitys["latitude"], js_activitys["longitude"], checkMsg = lat_long(data_activity["lat1"],
                                                                                             data_activity["lat2"],
                                                                                             data_activity["lat3"],
                                                                                             data_activity["long1"],
                                                                                             data_activity["long2"],
                                                                                             data_activity["long3"])

                    returnMsg = "Position manquante (latitude / longitude): Le " + str(data_activity["date"]).replace(
                        " 00:00:00", " à ") + str(data_activity["heure"])
                    if checkMsg:
                        allMessages.append(returnMsg)
                else:
                    js_activitys["latitude"], js_activitys["longitude"] = get_lat_long(allData,
                                                                                       info_bat['Depart_Port'])

            elif ( (index == last) or ("RTP" in data_activity["type_declaration"].upper()) ) :
                if data_activity["lat1"] != None and data_activity["lat2"] != None and \
                        data_activity["lat3"] != None and data_activity["long1"] != None and \
                        data_activity["long2"] != None and data_activity["long3"] != None:
                    js_activitys["latitude"], js_activitys["longitude"], checkMsg = lat_long(data_activity["lat1"],
                                                                                             data_activity["lat2"],
                                                                                             data_activity["lat3"],
                                                                                             data_activity["long1"],
                                                                                             data_activity["long2"],
                                                                                             data_activity["long3"])

                    returnMsg = "Position manquant (latitude / longitude): Le " + str(data_activity["date"]).replace(
                        " 00:00:00", " à ") + str(data_activity["heure"])
                    if checkMsg:
                        allMessages.append(returnMsg)
                else:
                    js_activitys["latitude"], js_activitys["longitude"] = get_lat_long(allData,
                                                                                       info_bat['Arrivee_Port'])
            else:
                js_activitys["latitude"], js_activitys["longitude"], checkMsg = lat_long(data_activity["lat1"], data_activity["lat2"],
                                                                                         data_activity["lat3"],
                                                                                         data_activity["long1"],
                                                                                         data_activity["long2"],
                                                                                         data_activity["long3"])

                returnMsg = "Position manquant (latitude / longitude): Le " + str(data_activity["date"]).replace(" 00:00:00",
                                                                                                        " à ") + str(
                    data_activity["heure"])
                if checkMsg:
                    allMessages.append(returnMsg)

            def schoolType(chaine, dico_code_sch_type):
                if chaine != None and chaine != "":
                    if ("libre" in chaine.lower()):
                        ## Table SchoolType code 2
                        return dico_code_sch_type["2"]

                    elif ("objet" in chaine.lower()):
                        ## Table SchoolType code 1
                        return dico_code_sch_type["1"]

                    elif ("ind" in chaine.lower()):
                        ## Table SchoolType code 0
                        return dico_code_sch_type["0"]

            def setCo_setSuc_vess(setCount, setSuccessStatus, vesselActivity):
                return setCount, setSuccessStatus, vesselActivity

            if data_activity["calee_type"] is not None:
                if ( ("FAR -" in data_activity["type_declaration"].upper()) and ("libre" in data_activity["calee_type"].lower()) ):
                    # Code 6
                    js_activitys["setCount"], js_activitys["setSuccessStatus"], js_activitys["vesselActivity"] = setCo_setSuc_vess(1, dico_code_setSucc["1"], vers_code_6)
                    js_activitys["schoolType"] = schoolType(data_activity["calee_type"], dico_code_sch_type)

                elif ( ("FAR -" in data_activity["type_declaration"].upper()) and ("objet" in data_activity["calee_type"].lower()) ):
                    # Code 6
                    js_activitys["setCount"], js_activitys["setSuccessStatus"], js_activitys["vesselActivity"] = setCo_setSuc_vess(1, dico_code_setSucc["0"], vers_code_6)
                    js_activitys["schoolType"] = schoolType(data_activity["calee_type"], dico_code_sch_type)

            elif ( ("FAR -" in data_activity["type_declaration"].upper()) and (data_activity["calee_type"] is None) ):
                # Code 13
                js_activitys["setCount"], js_activitys["setSuccessStatus"], js_activitys["vesselActivity"] = setCo_setSuc_vess(0, None, vers_code_13)
                js_activitys["schoolType"] = schoolType(data_activity["calee_type"], dico_code_sch_type)

            elif ( ("DEP -" in data_activity["type_declaration"].upper()) or ("RTP -" in data_activity["type_declaration"].upper()) ):
                # Code 0
                js_activitys["setCount"], js_activitys["setSuccessStatus"], js_activitys["vesselActivity"] = setCo_setSuc_vess(None, None, vers_code_0)
                js_activitys["schoolType"] = None

            elif ( "FAR0 -" in data_activity["type_declaration"].upper() ):
                # Code 103
                js_activitys["setCount"], js_activitys["setSuccessStatus"], js_activitys["vesselActivity"] = setCo_setSuc_vess(0, None, vers_code_103)
                js_activitys["schoolType"] = None

            js_activitys["informationSource"] = id_infoSource
            js_activitys["dataQuality"] = id_dataQua
            js_activitys["currentFpaZone"], comment_temp = fpaZone_id(data_activity["zee"], tab_fpa, allData)
            if (comment_temp != "") and (comment_temp != None):
                js_activitys["comment"] = js_activitys["comment"] + " # " + comment_temp

            ObserSys0 = getSome(allData, argment="code=0", moduleName="ObservedSystem")[0]["topiaId"]
            observedSystem = [ObserSys0]


            js_activitys["wind"] = get_wind_id_interval(allData, "Wind", data_activity["vent_vit"])

            nb_lignes = len(vals[1])  # Nombre total de lignes dans vals[1]

            if nb_lignes >= 2:
                dernier_element        = vals[1].iloc[-1]  # Récupération de le dernier élément de datas
                fpa_suiv               = dernier_element["zee"]
                type_decla_suiv        = dernier_element["type_declaration"].lower()[:3]
                comment_suiv           = dernier_element["type_declaration"]

            if ( (type_decla_suiv == "cox") and ("coe" == data_activity["type_declaration"].lower()[:3]) or (type_decla_suiv == "coe") and ("cox" == data_activity["type_declaration"].lower()[:3]) or (type_decla_suiv == "coe") and ("coe" == data_activity["type_declaration"].lower()[:3]) or (type_decla_suiv == "cox") and ("cox" == data_activity["type_declaration"].lower()[:3]) ):

                data_activity   = vals[1].iloc[-1]  # Récupération de le dernier élément de datas

                # faire le changement de zone
                # js_activitys["number"] = nb_prece
                js_activitys["previousFpaZone"], comment_temp = fpaZone_id(data_activity["zee"], tab_fpa, allData)
                js_activitys["nextFpaZone"], comment_temp = fpaZone_id(fpa_suiv, tab_fpa, allData)
                js_activitys["currentFpaZone"] = None

                if (comment_prece is not None) and ((js_activitys["comment"] is not None) and (js_activitys["comment"] != "Aucun commentaire")):
                    js_activitys["comment"] = js_activitys["comment"] + " ==> " + comment_suiv
                if (comment_prece is not None):
                    js_activitys["comment"] = data_activity["type_declaration"] + " ==> " + comment_suiv
                elif (comment_prece is None) and (js_activitys["comment"] is not None):
                    js_activitys["comment"] = js_activitys["comment"]
                elif (comment_prece is not None) and ((js_activitys["comment"] is None) or (js_activitys["comment"] == "Aucun commentaire")):
                    js_activitys["comment"] = data_activity["type_declaration"]
                else:
                    js_activitys["comment"] = "Aucun commentaire"

                if (comment_temp != "") and (comment_temp != None):
                    js_activitys["comment"] = js_activitys["comment"] + " # " + comment_temp

                js_activitys["vesselActivity"] = vers_code_21


                js_activitys["latitude"], js_activitys["longitude"], checkMsg = lat_long(data_activity["lat1"], data_activity["lat2"],
                                                                                         data_activity["lat3"],
                                                                                         data_activity["long1"],
                                                                                         data_activity["long2"],
                                                                                         data_activity["long3"])

                # ajouter la nouvelle activité
                activite.append(js_activitys)

                # Reinitialliser les variables
                nb = nb_prece

            else:
                activite.append(js_activitys)


            js_activitys["number"] = int(number)
            Som_thon                = 0
            nb += 1
            number += 1

        js_routeLogbooks = js_routeLogbook(activite)

        js_routeLogbooks["date"] = str(date).replace(" ", "T") + ".000Z"

        routes.append(js_routeLogbooks)

        activite = []

        nb = 1

    js_contents = js_content(routes, oce, prg)
    if not_time:
        js_contents["activitiesAcquisitionMode"] = "BY_NUMBER"
    else:
        js_contents["activitiesAcquisitionMode"] = "BY_TIME"

    # noinspection PyBroadException
    try:
        # si plusieurs Rechercher celui qui a le code le plus elévé avec toujours son status == 1
        js_contents["vessel"] = getId(allData, "Vessel",
                                      argment="label2=" + info_bat['Navire'] + "&filters.status=enabled", nbArg=3)
        if js_contents["vessel"] is None:
            js_contents["vessel"] = getId(allData, "Vessel",
                                          argment="label2=" + info_bat['Navire'].upper() + "&filters.status=enabled",
                                          nbArg=3)
        if js_contents["vessel"] is None:
            js_contents["vessel"] = getId(allData, "Vessel",
                                          argment="label2=" + info_bat['Navire'].lower() + "&filters.status=enabled",
                                          nbArg=3)
        if js_contents["vessel"] is None:
            allMessages.append(
                "Le nom du Navire n'a pas été trouvé. Veuillez verifier son existance dans la base de données puis corriger dans le livre de bord.")
    except:
        pass

    if info_bat['Depart_Port'] == None:
        js_contents["departureHarbour"] = getId(allData, "Harbour", argment="code=999")
    else:
        resu = getId(allData, "Harbour", argment="label2=" + (info_bat['Depart_Port']).upper())
        if resu == None:
            js_contents["departureHarbour"] = getId(allData, "Harbour", argment="code=999")
        else:
            js_contents["departureHarbour"] = resu

    if info_bat['Arrivee_Port'] == None:
        js_contents["landingHarbour"] = getId(allData, "Harbour", argment="code=999")
    else:
        resu = getId(allData, "Harbour", argment="label2=" + (info_bat['Arrivee_Port']).upper())
        if resu == None:
            js_contents["landingHarbour"] = getId(allData, "Harbour", argment="code=999")
        else:
            js_contents["landingHarbour"] = resu

    if info_bat['Depart_Date'] == None:
        js_contents["startDate"] = None
    else:
        js_contents["startDate"] = info_bat['Depart_Date'] + "T00:00:00.000Z"  # "2021-03-02T00:00:00.000Z" #

    if info_bat['Arrivee_Date'] == None:
        js_contents["endDate"] = None
    else:
        js_contents["endDate"] = info_bat['Arrivee_Date'] + "T00:00:00.000Z"


    js_contents["captain"], js_contents["logbookDataEntryOperator"] = cap_obs_sea(allData, info_bat)

    js_contents["loch"] = info_bat['Arrivee_Loch']
    js_contents["homeId"] = str(info_bat["mar_homeId"])
    js_contents["observationsProgram"] = None
    js_contents['logbookComment'] = "NB: Service Web"

    return allMessages, js_contents